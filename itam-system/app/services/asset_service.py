import csv
import re
from zipfile import BadZipFile
from datetime import datetime
from io import BytesIO, StringIO
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import and_, func, or_, text
from sqlalchemy.orm import Session

from app.core.time import app_datetime_to_utc, app_end_of_day, app_now, format_app_datetime, utc_now
from app.models.asset import Asset
from app.models.audit_response import AuditResponse
from app.models.audit_log import AssetChangeLog
from app.models.checkout import AssetCheckout
from app.models.company import Company
from app.models.file import AssetAttachment
from app.models.inventory import InventoryComponentInstallation, InventoryLedger, InventoryLicenseSeat, InventoryLicenseSeatHistory
from app.models.lifecycle import Lifecycle
from app.models.product import DeviceType, ProductCatalog
from app.models.repair import RepairRecord
from app.models.scan_binding import AssetScanBinding
from app.models.scrap import ScrapRequest
from app.models.stocktake import StocktakeItem, StocktakeScanLog
from app.models.user import UserDirectory
from app.schemas.asset import AssetBatchCheckinCreate, AssetBatchCheckoutCreate, AssetBatchImport, AssetBatchUpdateCreate, AssetCheckinCreate, AssetCheckoutCreate, AssetCreate, AssetImportRow, AssetTextImport, AssetUpdate
from app.services.audit_log_service import AuditLogService
from app.services.asset_residual_service import AssetResidualService
from app.services.lifecycle_service import LifecycleService
from app.services.notification_service import NotificationService
from app.services.number_service import NumberService
from app.services.supplier_service import SupplierService
from app.core.security import can_view_all_data, is_department_manager, scoped_dept_id


class AssetValidationError(ValueError):
    pass


class AssetService:
    DEFAULT_COMPANY = "未设置公司"
    ASSIGNED_STATUSES = {"in_use", "borrowed"}
    UNASSIGNED_STATUSES = {
        "pending_acceptance", "in_stock", "idle", "ready_scrap", "pending_scrap",
        "scrapped", "disposed", "lost",
    }
    WORKFLOW_STATUSES = {"pending_acceptance", "pending_scrap", "scrapped", "disposed", "lost"}
    TERMINAL_STATUSES = {"scrapped", "disposed", "lost"}
    CHECKOUT_ALLOWED_FROM = {"in_stock", "idle"}
    CHECKIN_ALLOWED_FROM = {"in_use", "borrowed", "out_stock", "repair"}
    VALID_STATUSES = {
        "pending_acceptance", "in_stock", "idle", "in_use",
        "borrowed", "out_stock", "repair", "ready_scrap", "pending_scrap",
        "scrapped", "disposed", "lost",
    }
    STATUS_TRANSITIONS = {
        "in_stock": {"idle", "in_use", "borrowed", "out_stock", "repair", "ready_scrap", "pending_scrap", "lost"},
        "idle": {"in_stock", "in_use", "borrowed", "out_stock", "repair", "ready_scrap", "pending_scrap", "lost"},
        "in_use": {"in_stock", "repair", "ready_scrap", "pending_scrap", "lost"},
        "borrowed": {"in_stock", "repair", "ready_scrap", "pending_scrap", "lost"},
        "out_stock": {"in_stock", "repair", "ready_scrap", "pending_scrap", "lost"},
        "repair": {"in_stock", "ready_scrap", "pending_scrap", "lost"},
        "ready_scrap": {"in_stock", "scrapped", "lost"},
        "pending_scrap": {"ready_scrap", "scrapped", "disposed", "lost"},
        "scrapped": {"disposed", "lost"},
        "disposed": set(),
        "lost": set(),
    }
    IMPORT_TEMPLATE_HEADERS = [
        "asset_id",
        "asset_no",
        "name",
        "category",
        "brand",
        "model",
        "sn",
        "purchase_price",
        "purchase_date",
        "purchase_approval_no",
        "purchase_supplier_name",
        "warranty_years",
        "status",
        "owner_user_id",
        "dept_id",
        "location",
        "company",
        "spec",
        "payment_time",
        "payment_no",
        "remark",
        "scan_codes",
        "status_time",
        "borrow_due_date",
        "disposal_method",
        "retirement_approval_no",
        "dispose_recipient_name",
    ]
    CHANGE_FIELD_LABELS = {
        "asset_id": "资产编号",
        "asset_no": "标签编号",
        "company": "所属公司",
        "name": "资产名称",
        "category": "设备类型",
        "brand": "品牌",
        "model": "型号",
        "sn": "序列号",
        "config": "规格配置",
        "purchase_price": "采购价格",
        "purchase_date": "采购日期",
        "purchase_approval_no": "采购审批单号",
        "purchase_supplier_name": "采购供应商",
        "warranty_expire_date": "质保到期",
        "warranty_months": "质保月数",
        "status": "状态",
        "owner_user_id": "责任人",
        "dept_id": "部门",
        "location": "位置",
        "remark": "备注",
    }

    @staticmethod
    def normalize_company(value: str | None) -> str | None:
        clean = (value or "").strip()
        return None if not clean or clean == AssetService.DEFAULT_COMPANY else clean

    @staticmethod
    def normalize_blank(value: str | None) -> str:
        return (value or "").strip()

    @staticmethod
    def normalize_asset_no(value: str | None, fallback: str | None = None) -> str:
        clean = AssetService.normalize_blank(value) or AssetService.normalize_blank(fallback)
        if not clean:
            raise AssetValidationError("标签编号不能为空")
        if clean == "0" or (clean.isdigit() and int(clean) == 0):
            raise AssetValidationError("标签编号不能为 0")
        return clean

    @staticmethod
    def validate_asset_identity_unique(db: Session, *, asset_no: str | None = None, sn: str | None = None, current_asset_id: str | None = None) -> None:
        if asset_no:
            query = db.query(Asset).filter(Asset.asset_no == asset_no)
            if current_asset_id:
                query = query.filter(Asset.asset_id != current_asset_id)
            if query.first():
                raise AssetValidationError(f"标签编号已存在：{asset_no}")
        if sn:
            query = db.query(Asset).filter(Asset.sn == sn)
            if current_asset_id:
                query = query.filter(Asset.asset_id != current_asset_id)
            if query.first():
                raise AssetValidationError(f"序列号已存在：{sn}")

    @staticmethod
    def validate_status_owner(asset: Asset, *, status_changed: bool = True, allow_workflow_statuses: bool = False) -> None:
        status = asset.status
        if status not in AssetService.VALID_STATUSES:
            raise AssetValidationError(f"unsupported asset status: {status}")
        has_owner = bool(AssetService.normalize_blank(asset.owner_user_id))
        has_location = bool(AssetService.normalize_blank(asset.location))
        if status_changed and status in AssetService.WORKFLOW_STATUSES and not allow_workflow_statuses:
            raise AssetValidationError("待验收、待处置登记、已报废、已丢失状态由流程控制，不能通过导入或手工状态变更直接设置")
        if status in AssetService.UNASSIGNED_STATUSES and has_owner:
            raise AssetValidationError("待验收、在库、闲置、待报废状态不能填写使用人/责任人；请清空使用人，或把状态改为 in_use、borrowed、out_stock")
        if status in AssetService.ASSIGNED_STATUSES and not has_owner:
            raise AssetValidationError("在用、借出状态必须填写使用人/责任人")
        if status == "out_stock" and not has_owner and not has_location:
            raise AssetValidationError("已出库状态必须填写领用人或出库地址")

    @staticmethod
    def ensure_asset_operable(asset: Asset, action: str = "操作") -> None:
        if asset.status == "scrapped":
            raise AssetValidationError(f"资产已报废，只能进行处置归档，不能执行{action}")
        if asset.status == "disposed":
            raise AssetValidationError(f"资产已处置归档，不能执行{action}")
        if asset.status == "lost":
            raise AssetValidationError(f"资产已丢失，不能执行{action}")

    @staticmethod
    def apply_warranty_expire(asset: Asset) -> None:
        if not asset.purchase_date or not asset.warranty_months:
            return
        asset.warranty_expire_date = AssetService.add_months(asset.purchase_date, asset.warranty_months)

    @staticmethod
    def add_months(value: datetime, months: int) -> datetime:
        month_index = value.month - 1 + int(months)
        year = value.year + month_index // 12
        month = month_index % 12 + 1
        day = min(value.day, AssetService.days_in_month(year, month))
        return value.replace(year=year, month=month, day=day)

    @staticmethod
    def days_in_month(year: int, month: int) -> int:
        if month == 12:
            next_month = datetime(year + 1, 1, 1)
        else:
            next_month = datetime(year, month + 1, 1)
        return (next_month - datetime(year, month, 1)).days

    @staticmethod
    def generate_asset_id(db: Session, prefix: str = "ITAM") -> str:
        return NumberService.next(db, f"asset:{prefix}", f"{prefix}-", 6)

    @staticmethod
    def create_asset(db: Session, payload: AssetCreate, operator: str = "system", user_context: dict | None = None) -> Asset:
        user = AssetService.find_user(db, payload.owner_user_id)
        AssetService.ensure_writable_department(
            (user.dept_id or user.dept_name) if user else payload.dept_id, user_context
        )
        asset_id = getattr(payload, "asset_id", None) or AssetService.generate_asset_id(db)
        asset_no = AssetService.normalize_asset_no(payload.asset_no, asset_id)
        sn = AssetService.normalize_blank(payload.sn) or None
        AssetService.validate_asset_identity_unique(db, asset_no=asset_no, sn=sn)
        asset = Asset(
            asset_id=asset_id,
            asset_no=asset_no,
            company=AssetService.normalize_company(payload.company),
            name=payload.name,
            category=payload.category,
            brand=payload.brand,
            model=payload.model,
            sn=sn,
            config=payload.config,
            purchase_price=payload.purchase_price,
            purchase_date=payload.purchase_date,
            purchase_approval_no=payload.purchase_approval_no,
            purchase_supplier_name=payload.purchase_supplier_name,
            warranty_expire_date=payload.warranty_expire_date,
            warranty_months=payload.warranty_months,
            status=payload.status,
            owner_user_id=user.user_id if user else payload.owner_user_id,
            dept_id=(user.dept_id or user.dept_name) if user else payload.dept_id,
            location=payload.location,
            remark=payload.remark,
        )
        AssetService.apply_warranty_expire(asset)
        AssetService.validate_status_owner(asset, status_changed=True)
        SupplierService.ensure_supplier(db, asset.purchase_supplier_name)
        db.add(asset)
        db.flush()
        LifecycleService.record(db, asset.asset_id, "CREATE", None, asset.status, operator)
        db.commit()
        db.refresh(asset)
        return AssetService.to_out(asset, user, db)

    @staticmethod
    def import_assets(db: Session, payload: AssetBatchImport, user_context: dict | None = None) -> dict:
        created_assets: list[Asset] = []
        updated_assets: list[Asset] = []
        errors: list[dict] = []
        skipped = 0
        scan_bindings_created = 0
        checkout_records_created = 0
        repair_records_created = 0
        scrap_requests_created = 0

        for index, row in enumerate(payload.items, start=1):
            row_scan_bindings_created = 0
            row_checkout_records_created = 0
            row_repair_records_created = 0
            row_scrap_requests_created = 0
            try:
                with db.begin_nested():
                    normalized = AssetService.normalize_import_row(row)
                    asset_id = normalized.asset_id or AssetService.generate_asset_id(db)
                    asset_no = AssetService.normalize_asset_no(normalized.asset_no, asset_id)
                    asset = db.get(Asset, asset_id)
                    if not asset and payload.overwrite:
                        asset = db.query(Asset).filter(Asset.asset_no == asset_no).first()
                    if asset:
                        AssetService.ensure_writable_department(asset.dept_id, user_context)
                    import_user = AssetService.find_user(db, normalized.owner_user_id)
                    AssetService.ensure_writable_department(
                        (import_user.dept_id or import_user.dept_name) if import_user
                        else normalized.dept_id or (asset.dept_id if asset else None),
                        user_context,
                    )
                    if asset and asset.asset_id != asset_id:
                        AssetService.rename_asset_id(db, asset, asset.asset_id, asset_id, payload.operator)
                    if asset and not payload.overwrite:
                        skipped += 1
                        errors.append({"row": index, "message": f"duplicate asset_id: {asset_id}", "data": row.model_dump()})
                        continue
                    AssetService.validate_asset_identity_unique(db, asset_no=asset_no, sn=normalized.sn, current_asset_id=asset.asset_id if asset else None)
                    AssetService.validate_status_owner(
                        SimpleNamespace(status=normalized.status, owner_user_id=normalized.owner_user_id, location=normalized.location),
                        allow_workflow_statuses=True,
                    )

                    old_status = asset.status if asset else None
                    if not asset:
                        asset = Asset(asset_id=asset_id)
                        db.add(asset)
                    asset.asset_no = asset_no
                    asset.company = AssetService.normalize_company(normalized.company)
                    asset.name = normalized.name
                    asset.category = normalized.category
                    asset.brand = normalized.brand
                    asset.model = normalized.model
                    asset.sn = normalized.sn
                    asset.config = normalized.config
                    asset.purchase_price = normalized.purchase_price
                    asset.purchase_date = normalized.purchase_date
                    asset.purchase_approval_no = normalized.purchase_approval_no
                    asset.purchase_supplier_name = normalized.purchase_supplier_name
                    asset.warranty_expire_date = normalized.warranty_expire_date
                    asset.warranty_months = normalized.warranty_months
                    asset.status = normalized.status
                    asset.owner_user_id = normalized.owner_user_id
                    asset.dept_id = normalized.dept_id
                    asset.location = normalized.location
                    asset.remark = normalized.remark
                    AssetService.apply_warranty_expire(asset)
                    AssetService.sync_owner_department(db, asset)
                    SupplierService.ensure_supplier(db, asset.purchase_supplier_name)
                    AssetService.ensure_company_from_import(db, normalized.company)
                    AssetService.ensure_product_catalog_from_import(db, normalized)
                    db.flush()
                    LifecycleService.record(db, asset.asset_id, "BATCH_IMPORT", old_status, asset.status, payload.operator)
                    row_scan_bindings_created = AssetService.sync_import_scan_bindings(
                        db,
                        asset.asset_id,
                        normalized.scan_codes,
                        payload.operator,
                    )
                    workflow_counts = AssetService.ensure_import_status_workflow(
                        db,
                        asset,
                        normalized,
                        payload.operator,
                    )
                    row_checkout_records_created = workflow_counts["checkout"]
                    row_repair_records_created = workflow_counts["repair"]
                    row_scrap_requests_created = workflow_counts["scrap"]
                    if old_status is None:
                        created_assets.append(asset)
                    else:
                        updated_assets.append(asset)
                scan_bindings_created += row_scan_bindings_created
                checkout_records_created += row_checkout_records_created
                repair_records_created += row_repair_records_created
                scrap_requests_created += row_scrap_requests_created
            except SQLAlchemyError as exc:
                errors.append({"row": index, "message": f"数据库保存失败：{AssetService.db_error_message(exc)}", "data": row.model_dump()})
            except Exception as exc:
                errors.append({"row": index, "message": str(exc), "data": row.model_dump()})

        db.commit()
        for asset in [*created_assets, *updated_assets]:
            db.refresh(asset)
        residual_config = AssetResidualService.get_config(db)

        return {
            "created": len(created_assets),
            "updated": len(updated_assets),
            "skipped": skipped,
            "scan_bindings_created": scan_bindings_created,
            "checkout_records_created": checkout_records_created,
            "repair_records_created": repair_records_created,
            "scrap_requests_created": scrap_requests_created,
            "errors": errors,
            "assets": [AssetService.to_out(asset, db=db, residual_config=residual_config) for asset in [*created_assets, *updated_assets]],
        }

    @staticmethod
    def sync_import_scan_bindings(db: Session, asset_id: str, scan_codes: list[str], operator: str) -> int:
        from app.services.scan_binding_service import ScanBindingService

        created = 0
        for scan_raw in AssetService.normalize_scan_codes(scan_codes):
            scan_key = ScanBindingService.normalize_scan_key(scan_raw)
            row = db.query(AssetScanBinding).filter(AssetScanBinding.scan_key == scan_key).first()
            if row and row.status == "active" and row.asset_id != asset_id:
                raise AssetValidationError(f"二维码内容已绑定其他资产：{row.asset_id}")
            if row:
                was_active = row.status == "active"
                row.asset_id = asset_id
                row.scan_raw = scan_raw
                row.scan_type = "qrcode"
                row.status = "active"
                row.remark = AssetService.join_notes(row.remark, "资产批量导入")
                row.updated_at = utc_now()
                created += 0 if was_active else 1
                action = "update_scan_code" if was_active else "rebind_scan_code"
            else:
                row = AssetScanBinding(
                    asset_id=asset_id,
                    scan_key=scan_key,
                    scan_raw=scan_raw,
                    scan_type="qrcode",
                    status="active",
                    remark="资产批量导入",
                    created_by=operator,
                )
                db.add(row)
                created += 1
                action = "bind_scan_code"
            AuditLogService.record_operation(
                db,
                "asset",
                action,
                operator,
                "asset_scan_binding",
                asset_id,
                f"批量导入二维码绑定 {asset_id}",
                {"asset_id": asset_id, "scan_raw": scan_raw, "scan_type": "qrcode"},
            )
        return created

    @staticmethod
    def ensure_import_status_workflow(db: Session, asset: Asset, row: AssetImportRow, operator: str) -> dict[str, int]:
        counts = {"checkout": 0, "repair": 0, "scrap": 0}
        status_time = row.status_time or utc_now()

        if row.status in {"in_use", "borrowed", "out_stock"}:
            checkout = AssetService.open_checkout_for_asset(db, asset.asset_id)
            if not checkout:
                user = AssetService.find_user(db, asset.owner_user_id)
                checkout = AssetCheckout(
                    asset_id=asset.asset_id,
                    checkout_type=row.status,
                    assignee_user_id=AssetService.normalize_blank(asset.owner_user_id) or None,
                    assignee_name=AssetService.user_label(user, asset.owner_user_id) if asset.owner_user_id else None,
                    dept_id=asset.dept_id,
                    location=asset.location,
                    due_date=row.borrow_due_date if row.status == "borrowed" else None,
                    status="open",
                    checked_out_at=status_time,
                    checked_out_by=operator,
                    remark=AssetService.join_notes(row.remark, "资产批量导入自动生成状态流程"),
                )
                db.add(checkout)
                AuditLogService.record_operation(
                    db,
                    "asset",
                    "create_checkout_from_import",
                    operator,
                    "asset_checkout",
                    asset.asset_id,
                    f"批量导入自动创建{AssetService.status_label(row.status)}记录 {asset.asset_id}",
                    {"asset_id": asset.asset_id, "checkout_type": row.status, "status_time": status_time},
                )
                counts["checkout"] = 1

        if row.status == "repair":
            active_repair = (
                db.query(RepairRecord)
                .filter(RepairRecord.asset_id == asset.asset_id, RepairRecord.finish_time.is_(None))
                .first()
            )
            if not active_repair:
                year = app_now().year
                repair = RepairRecord(
                    repair_no=NumberService.next(db, f"repair:{year}", f"RP-{year}-", 4),
                    asset_id=asset.asset_id,
                    repair_time=status_time,
                    repair_type="导入维修",
                    fault_reason=row.remark or "批量导入同步的维修中资产，故障原因待补充",
                    repair_cost=0,
                    operator=operator,
                    status="维修中",
                    remark="由资产批量导入自动创建",
                )
                db.add(repair)
                AuditLogService.record_operation(
                    db,
                    "repair",
                    "create_from_import",
                    operator,
                    "repair",
                    repair.repair_no,
                    f"批量导入自动创建维修单 {asset.asset_id}",
                    {"asset_id": asset.asset_id, "status_time": status_time},
                )
                counts["repair"] = 1

        counts["scrap"] = AssetService.ensure_import_scrap_request(db, asset, row, status_time, operator)
        return counts

    @staticmethod
    def ensure_import_scrap_request(
        db: Session,
        asset: Asset,
        row: AssetImportRow,
        status_time: datetime,
        operator: str,
    ) -> int:
        status = row.status
        if status not in {"pending_scrap", "scrapped", "disposed"}:
            return 0
        existed = db.query(ScrapRequest).filter(ScrapRequest.asset_id == asset.asset_id).first()
        if existed:
            return 0

        year = app_now().year
        completed = status in {"scrapped", "disposed"}
        disposal_method = AssetService.normalize_blank(row.disposal_method)
        if status == "scrapped" and not disposal_method:
            disposal_method = "报废"
        request = ScrapRequest(
            request_no=NumberService.next(db, f"scrap:{year}", f"SC-{year}-", 4),
            retirement_flow_no=NumberService.next(db, f"retirement_flow:{year}", f"RT-{year}-", 4),
            asset_id=asset.asset_id,
            asset_name=asset.name,
            asset_sn=asset.sn,
            company=asset.company,
            category=asset.category,
            brand=asset.brand,
            model=asset.model,
            owner_user_id=asset.owner_user_id,
            dept_id=asset.dept_id,
            location=asset.location,
            purchase_price=asset.purchase_price,
            purchase_date=asset.purchase_date,
            purchase_approval_no=asset.purchase_approval_no,
            purchase_supplier_name=asset.purchase_supplier_name,
            applicant=operator,
            reason="资产批量导入时状态为待处置或已报废，自动创建处置登记",
            estimated_residual_value=AssetResidualService.calculate_asset(asset, db=db),
            final_residual_value=AssetResidualService.calculate_asset(asset, db=db) if completed else 0,
            disposal_method=disposal_method if completed else None,
            retirement_date=status_time if completed else None,
            retirement_approval_no=AssetService.normalize_blank(row.retirement_approval_no),
            disposal_remark=(
                AssetService.join_notes(row.remark, "历史已处置资产批量导入")
                if completed
                else None
            ),
            dispose_recipient_name=(
                AssetService.normalize_blank(row.dispose_recipient_name)
                if completed and disposal_method == "员工领用"
                else None
            ),
            disposed_by=operator if completed else None,
            disposed_at=status_time if completed else None,
            status="已处置" if completed else "待处置",
        )
        db.add(request)
        LifecycleService.record(
            db,
            asset.asset_id,
            "SCRAP_REQUEST",
            asset.status,
            asset.status,
            operator,
            LifecycleService.structured_remark(
                reason=request.reason,
                object=f"报废单 {request.request_no}",
                location=asset.location,
                extra={"retirement_flow_no": request.retirement_flow_no},
            ),
        )
        AuditLogService.record_operation(
            db,
            "scrap",
            "create_from_import",
            operator,
            "scrap_request",
            request.request_no,
            f"批量导入自动创建报废处置登记 {asset.asset_id}",
            {"asset_id": asset.asset_id, "status": status, "retirement_flow_no": request.retirement_flow_no},
        )
        return 1

    @staticmethod
    def db_error_message(exc: SQLAlchemyError) -> str:
        detail = str(getattr(exc, "orig", exc)).strip()
        if "Duplicate entry" in detail and "assets.sn" in detail:
            return "资产序列号已存在，请检查 SN 是否重复"
        if "Duplicate entry" in detail and "PRIMARY" in detail:
            return "资产编号已存在，请检查 asset_id 是否重复"
        if "Data too long" in detail:
            return "字段内容过长，请检查该行文本长度"
        return detail or "请检查导入数据是否符合要求"

    @staticmethod
    def import_assets_from_text(db: Session, payload: AssetTextImport, user_context: dict | None = None) -> dict:
        items = AssetService.parse_import_text(payload.content)
        return AssetService.import_assets(db, AssetBatchImport(operator=payload.operator, overwrite=payload.overwrite, items=items), user_context)

    @staticmethod
    def import_assets_from_excel(db: Session, content: bytes, operator: str = "asset-import", overwrite: bool = False, user_context: dict | None = None) -> dict:
        items = AssetService.parse_import_excel(content)
        return AssetService.import_assets(db, AssetBatchImport(operator=operator, overwrite=overwrite, items=items), user_context)

    @staticmethod
    def preview_import_assets(db: Session, items: list[AssetImportRow], overwrite: bool = False) -> dict:
        from app.services.scan_binding_service import ScanBindingService

        errors: list[dict] = []
        preview_items: list[dict] = []
        seen_sn: set[str] = set()
        seen_asset_id: set[str] = set()
        seen_asset_no: set[str] = set()
        seen_scan_keys: dict[str, str] = {}

        for index, row in enumerate(items, start=1):
            try:
                normalized = AssetService.normalize_import_row(row)
                asset_no = AssetService.normalize_asset_no(normalized.asset_no, normalized.asset_id or f"PREVIEW-{index}")
                if asset_no in seen_asset_no:
                    raise AssetValidationError(f"duplicate asset_no: {asset_no}")
                if not overwrite:
                    AssetService.validate_asset_identity_unique(db, asset_no=asset_no)
                seen_asset_no.add(asset_no)
                if normalized.sn:
                    if normalized.sn in seen_sn or (not overwrite and db.query(Asset).filter(Asset.sn == normalized.sn).first()):
                        raise AssetValidationError(f"duplicate sn: {normalized.sn}")
                    seen_sn.add(normalized.sn)
                if normalized.asset_id:
                    if normalized.asset_id in seen_asset_id or (not overwrite and db.get(Asset, normalized.asset_id)):
                        raise AssetValidationError(f"duplicate asset_id: {normalized.asset_id}")
                    seen_asset_id.add(normalized.asset_id)
                target_asset_id = normalized.asset_id
                if overwrite and not target_asset_id:
                    existing_asset = db.query(Asset).filter(Asset.asset_no == asset_no).first()
                    target_asset_id = existing_asset.asset_id if existing_asset else None
                preview_identity = target_asset_id or asset_no
                for scan_raw in normalized.scan_codes:
                    scan_key = ScanBindingService.normalize_scan_key(scan_raw)
                    previous_identity = seen_scan_keys.get(scan_key)
                    if previous_identity and previous_identity != preview_identity:
                        raise AssetValidationError(f"二维码内容在导入文件中重复：{scan_raw}")
                    seen_scan_keys[scan_key] = preview_identity
                    existing_binding = (
                        db.query(AssetScanBinding)
                        .filter(AssetScanBinding.scan_key == scan_key, AssetScanBinding.status == "active")
                        .first()
                    )
                    if existing_binding and existing_binding.asset_id != target_asset_id:
                        raise AssetValidationError(f"二维码内容已绑定其他资产：{existing_binding.asset_id}")
                AssetService.validate_status_owner(
                    SimpleNamespace(status=normalized.status, owner_user_id=normalized.owner_user_id, location=normalized.location),
                    allow_workflow_statuses=True,
                )
                preview_items.append({"row": index, "valid": True, "data": normalized.model_dump(mode="json")})
            except Exception as exc:
                errors.append({"row": index, "message": str(exc), "data": row.model_dump(mode="json")})
                preview_items.append({"row": index, "valid": False, "data": row.model_dump(mode="json")})

        return {
            "total": len(items),
            "valid": len([item for item in preview_items if item["valid"]]),
            "errors": errors,
            "items": preview_items,
        }

    @staticmethod
    def preview_import_text(db: Session, payload: AssetTextImport) -> dict:
        return AssetService.preview_import_assets(db, AssetService.parse_import_text(payload.content), overwrite=payload.overwrite)

    @staticmethod
    def preview_import_excel(db: Session, content: bytes, overwrite: bool = False) -> dict:
        return AssetService.preview_import_assets(db, AssetService.parse_import_excel(content), overwrite=overwrite)

    @staticmethod
    def build_import_template() -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "资产导入"
        sheet.append(AssetService.IMPORT_TEMPLATE_HEADERS)
        sheet.append(
            [
                "",
                "NB-001",
                "ThinkPad X1 Carbon",
                "笔记本电脑",
                "Lenovo",
                "X1 Carbon Gen 12",
                "SN-IMPORT-001",
                15000,
                "2026-06-24",
                "OA-20260624-001",
                "联想授权供应商",
                3,
                "in_stock",
                "",
                "IT",
                "上海IT仓",
                "总部",
                "32G/1TB",
                "",
                "",
                "关键岗位备用机",
                "https://asset.example/nb-001",
                "2026-06-24 09:00:00",
                "",
            ]
        )
        sheet.append(
            [
                "",
                "DP-001",
                "Dell U2723QE",
                "显示器",
                "Dell",
                "U2723QE",
                "SN-IMPORT-002",
                3999,
                "2026-06-24",
                "OA-20260624-001",
                "Dell渠道商",
                3,
                "in_use",
                "U-ADMIN",
                "IT",
                "上海办公区",
                "总部",
                "27英寸 4K",
                "",
                "",
                "设计部高色准显示器",
                "QR-DP-001；QR-DP-001-LEGACY",
                "2026-06-24 10:00:00",
                "2026-07-24 18:00:00",
            ]
        )

        example = workbook.create_sheet("填写示例")
        example.append(AssetService.IMPORT_TEMPLATE_HEADERS)
        for values in sheet.iter_rows(min_row=2, max_row=3, values_only=True):
            example.append(list(values))
        sheet.delete_rows(2, 2)

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for cell in example[1]:
            cell.fill = header_fill
            cell.font = header_font

        widths = {
            "A": 18,
            "B": 24,
            "C": 16,
            "D": 16,
            "E": 20,
            "F": 20,
            "G": 16,
            "H": 16,
            "I": 22,
            "J": 22,
            "K": 16,
            "L": 16,
            "M": 18,
            "N": 16,
            "O": 18,
            "P": 18,
            "Q": 20,
            "R": 18,
            "S": 18,
            "T": 22,
            "U": 28,
            "V": 36,
            "W": 22,
            "X": 22,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
            example.column_dimensions[column].width = width
        sheet.freeze_panes = "A2"
        example.freeze_panes = "A2"

        status_validation = DataValidation(
            type="list",
            formula1='"in_stock,in_use,idle,borrowed,out_stock,repair,ready_scrap,pending_scrap,scrapped,disposed,lost"',
            allow_blank=False,
        )
        sheet.add_data_validation(status_validation)
        status_validation.add("M2:M500")

        instruction = workbook.create_sheet("字段说明")
        instruction.append(["字段", "是否必填", "说明"])
        rows = [
            ("asset_id", "否", "资产编号；留空时系统自动生成"),
            ("asset_no", "否", "标签编号，可填写公司内部贴纸编码"),
            ("name", "是", "资产名称"),
            ("category", "是", "设备类型，如 笔记本电脑、显示器"),
            ("brand", "否", "品牌"),
            ("model", "否", "型号"),
            ("sn", "否", "序列号；重复序列号会跳过导入"),
            ("purchase_price", "否", "采购价格，数字"),
            ("purchase_date", "否", "采购日期，格式 YYYY-MM-DD"),
            ("purchase_approval_no", "否", "采购审批单号或采购单号"),
            ("purchase_supplier_name", "否", "采购供应商"),
            ("warranty_years", "否", "维保年限，系统会换算为月数并计算质保到期"),
            ("status", "是", "支持英文状态；也可填写在库、在用、待报废、待处置登记、已报废、已处置、已丢失等中文状态"),
            ("owner_user_id", "按状态", "in_use、borrowed 必填；out_stock 可填写领用人或出库地址；库存/闲置/待报废必须留空"),
            ("dept_id", "否", "部门编号或部门名称"),
            ("location", "否", "当前位置"),
            ("company", "否", "所属公司"),
            ("spec", "否", "规格配置"),
            ("payment_time", "否", "付款时间或财务入账时间"),
            ("payment_no", "否", "付款单号或财务凭证号"),
            ("remark", "否", "备注/特殊说明，例如备用机、涉密、借测、待补配件"),
            ("scan_codes", "否", "二维码扫码返回的原始内容；多个内容用换行、英文分号或中文分号分隔，导入后直接绑定资产"),
            ("status_time", "否", "当前状态发生时间，格式 YYYY-MM-DD HH:MM:SS；留空使用导入时间"),
            ("borrow_due_date", "借用状态", "计划归还时间，仅 borrowed 状态使用，格式 YYYY-MM-DD HH:MM:SS"),
            ("disposal_method", "已处置状态", "实际处置方式：报废、变卖、员工领用"),
            ("retirement_approval_no", "否", "退役审批单号或历史退役流程编号"),
            ("dispose_recipient_name", "员工领用处置", "历史员工领用处置的接收人姓名或账号"),
        ]
        for row in rows:
            instruction.append(row)
        for cell in instruction[1]:
            cell.fill = header_fill
            cell.font = header_font
        instruction.column_dimensions["A"].width = 24
        instruction.column_dimensions["B"].width = 12
        instruction.column_dimensions["C"].width = 74

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def parse_import_text(content: str) -> list[AssetImportRow]:
        cleaned = content.strip()
        if not cleaned:
            return []
        delimiter = "\t" if "\t" in cleaned.splitlines()[0] else ","
        reader = csv.DictReader(StringIO(cleaned), delimiter=delimiter)
        return [AssetService.row_from_mapping(row) for row in reader if any(row.values())]

    @staticmethod
    def parse_import_excel(content: bytes) -> list[AssetImportRow]:
        if not content:
            raise AssetValidationError("Excel 文件为空，请重新选择文件")
        try:
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except BadZipFile as exc:
            raise AssetValidationError("Excel 文件无法打开，可能不是有效的 .xlsx/.xlsm 文件或文件已损坏") from exc
        except InvalidFileException as exc:
            raise AssetValidationError("Excel 文件格式不支持，请上传 .xlsx 或 .xlsm 文件") from exc
        except Exception as exc:
            raise AssetValidationError(f"Excel 文件解析失败：{exc}") from exc

        sheet = workbook.active
        if not sheet:
            raise AssetValidationError("Excel 文件没有可读取的工作表")
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise AssetValidationError("Excel 工作表为空，请至少保留表头行")

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        if not any(headers):
            raise AssetValidationError("Excel 第一行没有表头，请使用导入模板或填写字段名")
        normalized_headers = {header.lower() for header in headers if header}
        name_headers = {"name", "product_name", "资产名称", "产品名称"}
        category_headers = {"category", "device_type", "设备类型", "类别"}
        missing = []
        if not normalized_headers & {value.lower() for value in name_headers}:
            missing.append("资产名称(name)")
        if not normalized_headers & {value.lower() for value in category_headers}:
            missing.append("设备类型(category)")
        if missing:
            raise AssetValidationError(f"Excel 表头缺少必填字段：{', '.join(missing)}")

        items: list[AssetImportRow] = []
        for values in rows[1:]:
            mapping = {
                headers[index]: value
                for index, value in enumerate(values)
                if index < len(headers) and headers[index] and value not in (None, "")
            }
            if mapping:
                items.append(AssetService.row_from_mapping(mapping))
        if not items:
            return []
        return items

    @staticmethod
    def normalize_scan_codes(value: Any) -> list[str]:
        values = value if isinstance(value, (list, tuple, set)) else re.split(r"[\r\n;；]+", str(value or ""))
        result: list[str] = []
        seen: set[str] = set()
        for item in values:
            clean = str(item or "").strip()
            normalized = " ".join(clean.lower().split())
            if not clean or normalized in seen:
                continue
            seen.add(normalized)
            result.append(clean)
        return result

    @staticmethod
    def row_from_mapping(row: dict[str, Any]) -> AssetImportRow:
        normalized_row = {str(key).strip(): value for key, value in row.items()}

        def pick(*keys: str, default=None):
            for key in keys:
                value = normalized_row.get(key)
                if value not in (None, ""):
                    return value
            return default

        price = pick("purchase_price", "price", "unit_price", "采购价格", "价格", "单价", default=0)
        warranty_years = AssetService.parse_int(pick("warranty_years", "维保年限", "质保年限"))
        warranty_months = warranty_years * 12 if warranty_years is not None else AssetService.parse_int(pick("warranty_months", "质保月数", "维保月数", "质保"))
        config = {
            "spec": pick("spec", "规格", "配置", default=""),
            "source": "batch_import",
        }
        payment_time = pick("payment_time")
        payment_no = pick("payment_no")
        if payment_time:
            config["payment_time"] = str(payment_time).strip()
        if payment_no:
            config["payment_no"] = str(payment_no).strip()
        return AssetImportRow(
            asset_id=pick("asset_id", "资产编号", "外部ID"),
            asset_no=pick("asset_no", "标签编号", "资产编码"),
            name=pick("name", "product_name", "产品名称", "资产名称", default="Unnamed Asset"),
            category=pick("category", "device_type", "设备类型", "类别", default="Other"),
            brand=pick("brand", "品牌"),
            model=pick("model", "型号"),
            sn=pick("sn", "serial_number", "序列号", "SN"),
            config=config,
            purchase_price=AssetService.parse_float(price, "采购价格"),
            purchase_date=AssetService.parse_datetime(pick("purchase_date", "采购日期", "采购时间")),
            purchase_approval_no=pick("purchase_approval_no", "采购审批单号", "审批单号", "采购单号"),
            purchase_supplier_name=pick("purchase_supplier_name", "采购供应商", "供应商"),
            warranty_expire_date=AssetService.parse_datetime(pick("warranty_expire_date", "质保到期", "质保到期日")),
            warranty_months=warranty_months,
            payment_time=payment_time,
            payment_no=payment_no,
            status=pick("status", "状态", default="in_stock"),
            owner_user_id=pick("owner_user_id", "owner", "使用人", "责任人"),
            dept_id=pick("dept_id", "dept", "部门"),
            location=pick("location", "位置"),
            company=pick("company", "公司", "所属公司"),
            remark=pick("remark", "备注", "特殊说明", "说明"),
            scan_codes=AssetService.normalize_scan_codes(
                pick("scan_codes", "qr_codes", "scan_raw", "二维码内容", "二维码", "扫码内容")
            ),
            status_time=AssetService.parse_datetime(
                pick("status_time", "状态时间", "状态发生时间", "流程时间")
            ),
            borrow_due_date=AssetService.parse_datetime(
                pick("borrow_due_date", "计划归还时间", "借用到期时间", "预计归还时间")
            ),
            disposal_method=pick("disposal_method", "实际处置方式", "处置方式", "处理手段"),
            retirement_approval_no=pick(
                "retirement_approval_no", "退役审批单号", "退役流程编号", "处置审批单号"
            ),
            dispose_recipient_name=pick(
                "dispose_recipient_name", "处置领用人", "员工领用人", "报废领用人"
            ),
        )

    @staticmethod
    def normalize_import_row(row: AssetImportRow) -> AssetImportRow:
        data = row.model_dump()
        if data.get("product_name") and not data.get("name"):
            data["name"] = data["product_name"]
        if data.get("owner") and not data.get("owner_user_id"):
            data["owner_user_id"] = data["owner"]
        if data.get("dept") and not data.get("dept_id"):
            data["dept_id"] = data["dept"]
        if data.get("price") is not None and not data.get("purchase_price"):
            data["purchase_price"] = data["price"]
        data["asset_id"] = AssetService.normalize_blank(data.get("asset_id")) or None
        data["asset_no"] = AssetService.normalize_blank(data.get("asset_no")) or None
        data["sn"] = AssetService.normalize_blank(data.get("sn")) or None
        status = AssetService.normalize_blank(data.get("status")) or "in_stock"
        data["status"] = {
            "在库": "in_stock",
            "库存": "in_stock",
            "闲置": "idle",
            "在用": "in_use",
            "借出": "borrowed",
            "已出库": "out_stock",
            "维修": "repair",
            "待报废": "ready_scrap",
            "待处置": "pending_scrap",
            "待处置登记": "pending_scrap",
            "报废": "scrapped",
            "已报废": "scrapped",
            "已处置": "disposed",
            "丢失": "lost",
            "已丢失": "lost",
        }.get(status, status)
        data["scan_codes"] = AssetService.normalize_scan_codes(data.get("scan_codes"))

        config = data.get("config") or {}
        if data.get("spec"):
            config["spec"] = data["spec"]
        if data.get("payment_time"):
            config["payment_time"] = str(data["payment_time"]).strip()
        if data.get("payment_no"):
            config["payment_no"] = str(data["payment_no"]).strip()
        config.pop("warehouse", None)
        data["config"] = config
        return AssetImportRow(**data)

    @staticmethod
    def ensure_company_from_import(db: Session, company: str | None) -> Company | None:
        name = AssetService.normalize_company(company)
        if not name:
            return None
        existed = db.query(Company).filter(Company.name == name).first()
        if existed:
            return existed
        item = Company(name=name, status="启用")
        db.add(item)
        return item

    @staticmethod
    def ensure_product_catalog_from_import(db: Session, row: AssetImportRow) -> ProductCatalog | None:
        product_name = AssetService.normalize_blank(row.name)
        device_type = AssetService.normalize_blank(row.category)
        if not product_name or not device_type:
            return None

        if not db.query(DeviceType).filter(DeviceType.name == device_type).first():
            db.add(DeviceType(name=device_type, description="由资产导入自动创建"))

        brand = AssetService.normalize_blank(row.brand)
        model = AssetService.normalize_blank(row.model)
        spec = AssetService.normalize_blank((row.config or {}).get("spec"))

        catalog = (
            db.query(ProductCatalog)
            .filter(func.lower(func.trim(ProductCatalog.product_name)) == product_name.lower())
            .order_by(ProductCatalog.id.asc())
            .first()
        )

        if not catalog:
            catalog = ProductCatalog(
                product_name=product_name,
                device_type=device_type,
                brand=brand or None,
                model=model or None,
                spec=spec or None,
                unit_price=row.purchase_price or 0,
                default_warehouse=AssetService.normalize_blank(row.location) or None,
                retirement_years=AssetService.parse_int((row.config or {}).get("retirement_years")),
            )
            db.add(catalog)
            return catalog

        if not catalog.brand and brand:
            catalog.brand = brand
        if not catalog.model and model:
            catalog.model = model
        if not catalog.spec and spec:
            catalog.spec = spec
        if not catalog.unit_price and row.purchase_price:
            catalog.unit_price = row.purchase_price
        if not catalog.default_warehouse and row.location:
            catalog.default_warehouse = row.location
        if not catalog.retirement_years:
            catalog.retirement_years = AssetService.parse_int((row.config or {}).get("retirement_years"))
        return catalog

    @staticmethod
    def filter_nullable_text(query, column, value: str):
        if value:
            return query.filter(column == value)
        return query.filter(or_(column.is_(None), column == ""))

    @staticmethod
    def list_assets(
        db: Session,
        page: int = 1,
        page_size: int = 0,
        keyword: str | None = None,
        status: str | None = None,
        category: str | None = None,
        company: str | None = None,
        supplier: str | None = None,
        user_context: dict | None = None,
        risk_filter: str | None = None,
        owner_user_id: str | None = None,
    ) -> dict:
        users = AssetService.users_by_identity(db)
        query = db.query(Asset)
        query = AssetService.apply_data_scope(query, user_context)
        clean_keyword = (keyword or "").strip()
        if clean_keyword:
            pattern = f"%{clean_keyword}%"
            query = query.filter(
                or_(
                    Asset.asset_id.like(pattern),
                    Asset.asset_no.like(pattern),
                    Asset.name.like(pattern),
                    Asset.dept_id.like(pattern),
                    Asset.sn.like(pattern),
                    Asset.brand.like(pattern),
                    Asset.model.like(pattern),
                    Asset.owner_user_id.like(pattern),
                    Asset.purchase_approval_no.like(pattern),
                    Asset.purchase_supplier_name.like(pattern),
                    Asset.remark.like(pattern),
                )
            )
        if status:
            query = query.filter(Asset.status == status)
        if category:
            query = query.filter(Asset.category == category)
        if supplier:
            query = query.filter(Asset.purchase_supplier_name == supplier)
        if company:
            stored = AssetService.normalize_company(company)
            query = query.filter(Asset.company.is_(None) if stored is None else Asset.company == stored)
        clean_owner_user_id = AssetService.normalize_blank(owner_user_id)
        if clean_owner_user_id:
            owner = AssetService.find_user(db, clean_owner_user_id)
            owner_identities = {clean_owner_user_id}
            if owner:
                owner_identities.update(
                    value
                    for value in [owner.user_id, owner.username, owner.external_id, owner.email, owner.display_name]
                    if value
                )
            normalized_identities = {
                str(value).strip().casefold()
                for value in owner_identities
                if str(value).strip()
            }
            owner_value = func.lower(func.trim(Asset.owner_user_id))
            identity_filters = [owner_value.in_(normalized_identities)]
            for identity in normalized_identities:
                for separator in ["-", "_", "/", "\\", "|", ",", ";", ":", "，", "；", "：", " "]:
                    identity_filters.extend([
                        owner_value.like(f"{identity}{separator}%"),
                        owner_value.like(f"%{separator}{identity}"),
                        owner_value.like(f"%{separator}{identity}{separator}%"),
                    ])
            query = query.filter(or_(*identity_filters))
        query = AssetService.apply_asset_risk_filter(query, risk_filter)

        total = query.count()
        query = query.order_by(Asset.created_at.desc())
        if page_size and page_size > 0:
            query = query.offset((max(page, 1) - 1) * page_size).limit(page_size)
        assets = query.all()
        rows = []
        offset_base = (max(page, 1) - 1) * page_size if page_size and page_size > 0 else 0
        residual_config = AssetResidualService.get_config(db)
        for index, asset in enumerate(assets):
            user = users.get(asset.owner_user_id or "")
            row = AssetService.to_out(asset, user, db, residual_config)
            row["display_id"] = max(total - offset_base - index, 1)
            # 展示层归一化：负责人/部门与用户目录保持一致，持久化同步由写路径的 sync_owner_department 负责
            if user:
                row["owner_user_id"] = user.user_id
                target_dept = user.dept_id or user.dept_name or asset.dept_id
                if target_dept:
                    row["dept_id"] = target_dept
            rows.append(row)
        return {"list": rows, "total": total, "page": max(page, 1), "page_size": page_size or total}

    @staticmethod
    def get_asset(db: Session, asset_id: str, user_context: dict | None = None) -> dict:
        asset = AssetService.get_scoped_asset(db, asset_id, user_context)
        user = AssetService.users_by_identity(db).get(asset.owner_user_id or "")
        row = AssetService.to_out(asset, user, db)
        row["display_id"] = AssetService.asset_display_id(db, asset)
        if user:
            row["owner_user_id"] = user.user_id
            target_dept = user.dept_id or user.dept_name or asset.dept_id
            if target_dept:
                row["dept_id"] = target_dept
        return row

    @staticmethod
    def asset_display_id(db: Session, asset: Asset) -> int:
        if not asset.created_at:
            return 1
        return (
            db.query(func.count(Asset.asset_id))
            .filter(
                or_(
                    Asset.created_at < asset.created_at,
                    and_(Asset.created_at == asset.created_at, Asset.asset_id <= asset.asset_id),
                )
            )
            .scalar()
            or 1
        )

    @staticmethod
    def apply_asset_risk_filter(query, risk_filter: str | None):
        clean_filter = (risk_filter or "").strip()
        if not clean_filter:
            return query
        active_status_filter = Asset.status.in_(["in_use", "borrowed", "out_stock"])
        warranty_overdue_filter = Asset.warranty_expire_date.isnot(None) & (Asset.warranty_expire_date < utc_now())
        retirement_overdue_filter = text(
            "purchase_date IS NOT NULL "
            "AND JSON_UNQUOTE(JSON_EXTRACT(config, '$.retirement_years')) IS NOT NULL "
            "AND CAST(JSON_UNQUOTE(JSON_EXTRACT(config, '$.retirement_years')) AS UNSIGNED) > 0 "
            "AND TIMESTAMPADD(YEAR, CAST(JSON_UNQUOTE(JSON_EXTRACT(config, '$.retirement_years')) AS UNSIGNED), purchase_date) < NOW()"
        )
        if clean_filter == "active_warranty_overdue":
            return query.filter(active_status_filter, warranty_overdue_filter)
        if clean_filter == "active_retirement_overdue":
            return query.filter(active_status_filter, retirement_overdue_filter)
        if clean_filter == "active_warranty_or_retirement_overdue":
            return query.filter(active_status_filter).filter(or_(warranty_overdue_filter, retirement_overdue_filter))
        return query

    @staticmethod
    def apply_data_scope(query, user_context: dict | None):
        user_context = user_context or {}
        role = (user_context.get("role") or "").lower()
        if role in {"admin", "auditor", "asset_manager"}:
            return query
        dept_id = user_context.get("dept_id") or user_context.get("dept_name")
        user_id = user_context.get("user_id")
        username = user_context.get("username")
        if role in {"dept_manager", "department_manager", "manager", "部门管理员"} and dept_id:
            return query.filter(Asset.dept_id == dept_id)
        identities = [value for value in [user_id, username] if value]
        if identities:
            return query.filter(Asset.owner_user_id.in_(identities))
        return query.filter(False)

    @staticmethod
    def ensure_writable_department(dept_id: str | None, user_context: dict | None) -> None:
        if not user_context or can_view_all_data(user_context):
            return
        if is_department_manager(user_context):
            current_dept = AssetService.normalize_blank(scoped_dept_id(user_context))
            target_dept = AssetService.normalize_blank(dept_id)
            if current_dept and target_dept == current_dept:
                return
            raise AssetValidationError("Department managers can only write assets in their department")
        raise AssetValidationError("Current account has no writable asset data scope")

    @staticmethod
    def get_scoped_asset(db: Session, asset_id: str, user_context: dict | None = None) -> Asset:
        asset = AssetService.apply_data_scope(
            db.query(Asset).filter(Asset.asset_id == asset_id), user_context
        ).first()
        if not asset:
            raise ValueError("asset not found")
        return asset

    @staticmethod
    def get_scoped_asset_for_update(db: Session, asset_id: str, user_context: dict | None = None) -> Asset:
        asset = AssetService.apply_data_scope(
            db.query(Asset).filter(Asset.asset_id == asset_id), user_context
        ).populate_existing().with_for_update().first()
        if not asset:
            raise ValueError("asset not found")
        return asset

    @staticmethod
    def asset_summary(db: Session, user_context: dict | None = None) -> dict:
        now = utc_now()
        current_month = datetime(now.year, now.month, 1)
        previous_month = datetime(now.year - 1, 12, 1) if now.month == 1 else datetime(now.year, now.month - 1, 1)

        scoped = AssetService.apply_data_scope(db.query(Asset), user_context)
        scoped_ids = scoped.with_entities(Asset.asset_id)
        total = scoped.count()
        total_value = scoped.with_entities(func.coalesce(func.sum(Asset.purchase_price), 0)).scalar() or 0
        managed_filter = or_(Asset.status.is_(None), ~Asset.status.in_(["scrapped", "disposed", "lost"]))
        managed_total = scoped.filter(managed_filter).count()
        managed_total_value = scoped.filter(managed_filter).with_entities(func.coalesce(func.sum(Asset.purchase_price), 0)).scalar() or 0
        current_month_count = scoped.filter(Asset.created_at >= current_month).count()
        previous_month_count = (
            scoped
            .filter(Asset.created_at >= previous_month, Asset.created_at < current_month)
            .count()
        )
        current_month_managed_count = (
            scoped
            .filter(managed_filter, Asset.created_at >= current_month)
            .count()
        )
        previous_month_managed_count = (
            scoped
            .filter(managed_filter, Asset.created_at >= previous_month, Asset.created_at < current_month)
            .count()
        )

        status_counts = {
            status or "unknown": count
            for status, count in db.query(Asset.status, func.count(Asset.asset_id)).filter(Asset.asset_id.in_(scoped_ids)).group_by(Asset.status).all()
        }
        category_counts = {
            category or "其他": count
            for category, count in db.query(Asset.category, func.count(Asset.asset_id)).filter(Asset.asset_id.in_(scoped_ids)).group_by(Asset.category).all()
        }
        managed_status_counts = {
            status or "unknown": count
            for status, count in db.query(Asset.status, func.count(Asset.asset_id)).filter(Asset.asset_id.in_(scoped_ids), managed_filter).group_by(Asset.status).all()
        }
        managed_category_counts = {
            category or "其他": count
            for category, count in db.query(Asset.category, func.count(Asset.asset_id)).filter(Asset.asset_id.in_(scoped_ids), managed_filter).group_by(Asset.category).all()
        }
        return {
            "total": total,
            "total_value": float(total_value),
            "managed_total": managed_total,
            "managed_total_value": float(managed_total_value),
            "status_counts": status_counts,
            "category_counts": category_counts,
            "managed_status_counts": managed_status_counts,
            "managed_category_counts": managed_category_counts,
            "current_month_count": current_month_count,
            "previous_month_count": previous_month_count,
            "current_month_managed_count": current_month_managed_count,
            "previous_month_managed_count": previous_month_managed_count,
        }

    @staticmethod
    def update_asset(db: Session, asset_id: str, payload: AssetUpdate, operator: str = "system", user_context: dict | None = None) -> Asset:
        asset = AssetService.apply_asset_update(db, asset_id, payload, operator, user_context)
        db.commit()
        db.refresh(asset)
        return AssetService.to_out(asset, db=db)

    @staticmethod
    def apply_asset_update(
        db: Session,
        asset_id: str,
        payload: AssetUpdate,
        operator: str = "system",
        user_context: dict | None = None,
        source: str = "asset_update",
    ) -> Asset:
        asset = AssetService.get_scoped_asset(db, asset_id, user_context)

        data = payload.model_dump(exclude_unset=True)
        if "status" in data and data["status"] != asset.status:
            raise AssetValidationError("asset status must be changed through checkout, checkin, repair, or scrap workflow")
        status_changed = "status" in data and data["status"] != asset.status
        owner_changed = (
            "owner_user_id" in data
            and AssetService.normalize_blank(data["owner_user_id"]) != AssetService.normalize_blank(asset.owner_user_id)
        )
        dept_changed = (
            "dept_id" in data
            and AssetService.normalize_blank(data["dept_id"]) != AssetService.normalize_blank(asset.dept_id)
        )
        location_changed = (
            "location" in data
            and AssetService.normalize_blank(data["location"]) != AssetService.normalize_blank(asset.location)
        )
        if asset.status in AssetService.TERMINAL_STATUSES and any(
            (status_changed, owner_changed, dept_changed, location_changed)
        ):
            raise AssetValidationError("已报废/已处置资产不能再修改状态、责任人、部门或位置")
        new_asset_id = AssetService.normalize_blank(data.pop("asset_id", None))
        original_values = AssetService.snapshot_asset(asset)
        if new_asset_id and new_asset_id != asset_id:
            AssetService.rename_asset_id(db, asset, asset_id, new_asset_id, operator)
        old_status = asset.status
        should_validate_status_owner = status_changed or owner_changed
        for key, value in data.items():
            if key == "asset_no":
                value = AssetService.normalize_asset_no(value)
                AssetService.validate_asset_identity_unique(db, asset_no=value, current_asset_id=asset.asset_id)
            if key == "sn":
                value = AssetService.normalize_blank(value) or None
                AssetService.validate_asset_identity_unique(db, sn=value, current_asset_id=asset.asset_id)
            if key == "company":
                value = AssetService.normalize_company(value)
            if key == "owner_user_id":
                value = AssetService.normalize_blank(value)
            setattr(asset, key, value)
        AssetService.apply_warranty_expire(asset)
        AssetService.sync_owner_department(db, asset)
        if should_validate_status_owner:
            AssetService.validate_status_owner(asset, status_changed=asset.status != old_status)
        SupplierService.ensure_supplier(db, asset.purchase_supplier_name)

        AssetService.record_asset_field_changes(db, asset, original_values, operator, source=source)
        AuditLogService.record_operation(
            db,
            module="asset",
            action="update",
            target_type="asset",
            target_id=asset.asset_id,
            operator=operator,
            summary=f"更新资产 {asset.asset_id}",
        )
        LifecycleService.record(db, asset.asset_id, "ASSET_UPDATE", old_status, asset.status, operator)
        return asset

    @staticmethod
    def batch_update_assets(db: Session, payload: AssetBatchUpdateCreate, operator: str = "system", user_context: dict | None = None) -> dict:
        rows: list[Asset] = []
        errors: list[dict] = []
        clean_ids = [AssetService.normalize_blank(asset_id) for asset_id in payload.asset_ids if AssetService.normalize_blank(asset_id)]
        unique_ids = list(dict.fromkeys(clean_ids))
        if not unique_ids:
            return {"success": 0, "failed": 0, "assets": [], "errors": []}
        try:
            for asset_id in unique_ids:
                rows.append(AssetService.apply_asset_update(db, asset_id, payload.updates, operator, user_context))
            db.commit()
            for asset in rows:
                db.refresh(asset)
            residual_config = AssetResidualService.get_config(db)
            return {"success": len(rows), "failed": 0, "assets": [AssetService.to_out(asset, db=db, residual_config=residual_config) for asset in rows], "errors": []}
        except (AssetValidationError, ValueError) as exc:
            db.rollback()
            errors = [{"asset_id": asset_id, "message": str(exc)} for asset_id in unique_ids]
            return {"success": 0, "failed": len(errors), "assets": [], "errors": errors}

    @staticmethod
    def snapshot_asset(asset: Asset) -> dict:
        return {field: getattr(asset, field, None) for field in AssetService.CHANGE_FIELD_LABELS if hasattr(asset, field)}

    @staticmethod
    def record_asset_field_changes(db: Session, asset: Asset, original: dict, operator: str, source: str) -> None:
        for field, old_value in original.items():
            new_value = getattr(asset, field, None)
            AuditLogService.record_asset_change(
                db,
                asset.asset_id,
                field,
                old_value,
                new_value,
                operator,
                AssetService.CHANGE_FIELD_LABELS.get(field, field),
                source,
            )

    @staticmethod
    def list_asset_changes(db: Session, asset_id: str, limit: int = 200, user_context: dict | None = None) -> list[dict]:
        AssetService.get_scoped_asset(db, asset_id, user_context)
        rows = (
            db.query(AssetChangeLog)
            .filter(AssetChangeLog.asset_id == asset_id)
            .order_by(AssetChangeLog.created_at.desc(), AssetChangeLog.id.desc())
            .limit(min(max(limit, 1), 500))
            .all()
        )
        return [
            {
                "id": row.id,
                "asset_id": row.asset_id,
                "field_name": row.field_name,
                "field_label": row.field_label or row.field_name,
                "old_value": row.old_value or "",
                "new_value": row.new_value or "",
                "operator": row.operator,
                "source": row.source,
                "created_at": row.created_at.isoformat(sep=" ", timespec="seconds") if row.created_at else "",
            }
            for row in rows
        ]

    @staticmethod
    def rename_asset_id(db: Session, asset: Asset, old_asset_id: str, new_asset_id: str, operator: str) -> None:
        if not new_asset_id:
            raise AssetValidationError("资产编号不能为空")
        if len(new_asset_id) > 64:
            raise AssetValidationError("资产编号不能超过 64 个字符")
        if db.get(Asset, new_asset_id):
            raise AssetValidationError(f"资产编号 {new_asset_id} 已存在，请换一个编号")

        mysql_fk_disabled = db.bind and db.bind.dialect.name in {"mysql", "mariadb"}
        if mysql_fk_disabled:
            db.execute(text("SET FOREIGN_KEY_CHECKS=0"))
        try:
            db.query(Lifecycle).filter(Lifecycle.asset_id == old_asset_id).update({Lifecycle.asset_id: new_asset_id}, synchronize_session=False)
            db.query(AssetAttachment).filter(AssetAttachment.asset_id == old_asset_id).update({AssetAttachment.asset_id: new_asset_id}, synchronize_session=False)
            db.query(RepairRecord).filter(RepairRecord.asset_id == old_asset_id).update({RepairRecord.asset_id: new_asset_id}, synchronize_session=False)
            db.query(ScrapRequest).filter(ScrapRequest.asset_id == old_asset_id).update({ScrapRequest.asset_id: new_asset_id}, synchronize_session=False)
            db.query(StocktakeItem).filter(StocktakeItem.asset_id == old_asset_id).update({StocktakeItem.asset_id: new_asset_id}, synchronize_session=False)
            db.query(StocktakeScanLog).filter(StocktakeScanLog.asset_id == old_asset_id).update({StocktakeScanLog.asset_id: new_asset_id}, synchronize_session=False)
            db.query(AssetCheckout).filter(AssetCheckout.asset_id == old_asset_id).update({AssetCheckout.asset_id: new_asset_id}, synchronize_session=False)
            db.query(AssetScanBinding).filter(AssetScanBinding.asset_id == old_asset_id).update({AssetScanBinding.asset_id: new_asset_id}, synchronize_session=False)
            db.query(InventoryLedger).filter(InventoryLedger.asset_id == old_asset_id).update({InventoryLedger.asset_id: new_asset_id}, synchronize_session=False)
            db.query(InventoryLicenseSeat).filter(InventoryLicenseSeat.asset_id == old_asset_id).update({InventoryLicenseSeat.asset_id: new_asset_id}, synchronize_session=False)
            db.query(InventoryLicenseSeatHistory).filter(InventoryLicenseSeatHistory.asset_id == old_asset_id).update({InventoryLicenseSeatHistory.asset_id: new_asset_id}, synchronize_session=False)
            db.query(InventoryComponentInstallation).filter(InventoryComponentInstallation.asset_id == old_asset_id).update({InventoryComponentInstallation.asset_id: new_asset_id}, synchronize_session=False)
            db.query(AssetChangeLog).filter(AssetChangeLog.asset_id == old_asset_id).update({AssetChangeLog.asset_id: new_asset_id}, synchronize_session=False)
            for response in db.query(AuditResponse).filter(AuditResponse.asset_id == old_asset_id).all():
                response.asset_id = new_asset_id
                response.violation_key = response.violation_key.replace(old_asset_id, new_asset_id, 1)
            asset.asset_id = new_asset_id
            db.flush()
        finally:
            if mysql_fk_disabled:
                db.execute(text("SET FOREIGN_KEY_CHECKS=1"))
        LifecycleService.record(
            db,
            new_asset_id,
            "ASSET_ID_CHANGE",
            None,
            asset.status,
            operator,
            LifecycleService.structured_remark(
                reason="资产编号调整",
                object=asset.asset_id,
                location=asset.location,
                extra={"old_asset_id": old_asset_id, "new_asset_id": new_asset_id},
            ),
        )

    @staticmethod
    def change_status(
        db: Session,
        asset_id: str,
        to_status: str,
        operator: str = "system",
        owner_user_id: str | None = None,
        dept_id: str | None = None,
        location: str | None = None,
        borrow_due_date: str | None = None,
        remark: str | None = None,
        user_context: dict | None = None,
        expected_from_statuses: set[str] | None = None,
    ) -> Asset:
        asset = AssetService.get_scoped_asset_for_update(db, asset_id, user_context)
        if expected_from_statuses is not None and asset.status not in expected_from_statuses:
            raise AssetValidationError(f"asset status changed concurrently: {asset.status}")
        from_status = asset.status
        previous_owner_user_id, previous_user, user = AssetService.apply_status_transition(
            db, asset, to_status, operator, owner_user_id, dept_id, location,
            borrow_due_date, remark,
        )
        db.commit()
        db.refresh(asset)
        AssetService.notify_status_change(db, asset, from_status, to_status, operator, previous_user, previous_owner_user_id, user)
        return AssetService.to_out(asset, user, db)

    @staticmethod
    def apply_status_transition(
        db: Session,
        asset: Asset,
        to_status: str,
        operator: str = "system",
        owner_user_id: str | None = None,
        dept_id: str | None = None,
        location: str | None = None,
        borrow_due_date: str | None = None,
        remark: str | None = None,
        event_type: str = "STATUS_CHANGE",
        record_lifecycle: bool = True,
        allow_workflow_statuses: bool = False,
    ) -> tuple[str | None, UserDirectory | None, UserDirectory | None]:
        from_status = asset.status
        AssetService.validate_transition(from_status, to_status)
        if from_status in AssetService.TERMINAL_STATUSES:
            raise AssetValidationError("已报废/已处置资产不能再做状态流转")
        previous_owner_user_id = asset.owner_user_id
        previous_user = AssetService.find_user(db, previous_owner_user_id)
        if owner_user_id is not None:
            asset.owner_user_id = AssetService.normalize_blank(owner_user_id)
        elif to_status in AssetService.UNASSIGNED_STATUSES:
            asset.owner_user_id = None
        user = AssetService.sync_owner_department(db, asset)
        if dept_id is not None and not user:
            asset.dept_id = dept_id
        if location is not None:
            asset.location = location
        asset.status = to_status
        AssetService.apply_borrow_due_date(asset, to_status, borrow_due_date)
        AssetService.validate_status_owner(
            asset,
            status_changed=to_status != from_status,
            allow_workflow_statuses=allow_workflow_statuses,
        )
        lifecycle_remark = AssetService.inventory_lifecycle_remark(
            to_status,
            remark,
            previous_owner_user_id,
            previous_user,
            asset.owner_user_id,
            user,
            asset.location,
            borrow_due_date,
        )
        if record_lifecycle:
            LifecycleService.record(db, asset.asset_id, event_type, from_status, to_status, operator, lifecycle_remark)
        AssetService.sync_checkout_record(
            db,
            asset,
            from_status,
            to_status,
            operator,
            previous_owner_user_id,
            previous_user,
            user,
            borrow_due_date,
            remark,
        )
        return previous_owner_user_id, previous_user, user

    @staticmethod
    def checkout_asset(db: Session, asset_id: str, payload: AssetCheckoutCreate, operator: str = "system", user_context: dict | None = None) -> Asset:
        asset = AssetService.get_scoped_asset(db, asset_id, user_context)
        AssetService.ensure_asset_operable(asset, "领用/出库")
        if asset.status not in AssetService.CHECKOUT_ALLOWED_FROM:
            raise AssetValidationError(f"当前状态为 {AssetService.status_label(asset.status)}，不能重复领用/出库；请先归还入库后再出库")
        checkout_type = payload.checkout_type or "in_use"
        if checkout_type not in {"in_use", "borrowed", "out_stock"}:
            raise AssetValidationError("领用类型只能是 in_use、borrowed 或 out_stock")
        if checkout_type in {"in_use", "borrowed"} and not AssetService.normalize_blank(payload.owner_user_id):
            raise AssetValidationError("领用/借出必须选择领用人")
        if checkout_type == "out_stock" and not AssetService.normalize_blank(payload.owner_user_id) and not AssetService.normalize_blank(payload.location):
            raise AssetValidationError("出库必须选择领用人或出库地址")
        return AssetService.change_status(
            db,
            asset_id,
            checkout_type,
            operator,
            payload.owner_user_id,
            payload.dept_id,
            payload.location,
            payload.due_date,
            payload.remark,
            user_context,
            expected_from_statuses=AssetService.CHECKOUT_ALLOWED_FROM,
        )

    @staticmethod
    def checkin_asset(db: Session, asset_id: str, payload: AssetCheckinCreate, operator: str = "system", user_context: dict | None = None) -> Asset:
        asset = AssetService.get_scoped_asset(db, asset_id, user_context)
        AssetService.ensure_asset_operable(asset, "归还/入库")
        if asset.status not in AssetService.CHECKIN_ALLOWED_FROM:
            raise AssetValidationError(f"当前状态为 {AssetService.status_label(asset.status)}，不能重复入库；只有在用、借出、已出库或维修中的资产可以入库")
        return AssetService.change_status(
            db,
            asset_id,
            "in_stock",
            operator,
            "",
            "",
            payload.location,
            None,
            payload.remark or "资产归还入库",
            user_context=user_context,
            expected_from_statuses=AssetService.CHECKIN_ALLOWED_FROM,
        )

    @staticmethod
    def validate_transition(from_status: str | None, to_status: str) -> None:
        if to_status not in AssetService.VALID_STATUSES:
            raise AssetValidationError(f"unsupported asset status: {to_status}")
        if from_status == to_status:
            return
        if to_status not in AssetService.STATUS_TRANSITIONS.get(from_status or "", set()):
            raise AssetValidationError(f"invalid asset status transition: {from_status or '-'} -> {to_status}")

    @staticmethod
    def apply_borrow_due_date(asset: Asset, to_status: str, borrow_due_date: str | None) -> None:
        config = dict(asset.config or {})
        clean_due_date = AssetService.normalize_blank(borrow_due_date)
        if to_status == "borrowed":
            if not clean_due_date:
                raise AssetValidationError("借出资产必须填写借用到期时间")
            config["borrow_due_date"] = clean_due_date
        else:
            config.pop("borrow_due_date", None)
        asset.config = config

    @staticmethod
    def sync_checkout_record(
        db: Session,
        asset: Asset,
        from_status: str | None,
        to_status: str,
        operator: str,
        previous_owner_user_id: str | None,
        previous_user: UserDirectory | None,
        current_user: UserDirectory | None,
        borrow_due_date: str | None,
        remark: str | None,
    ) -> None:
        if to_status in AssetService.UNASSIGNED_STATUSES:
            AssetService.close_open_checkout(db, asset, operator, asset.location, remark or "资产归还入库")
            return
        if to_status not in {"in_use", "borrowed", "out_stock"}:
            return

        open_checkout = AssetService.open_checkout_for_asset(db, asset.asset_id)
        same_assignment = (
            open_checkout
            and open_checkout.checkout_type == to_status
            and (open_checkout.assignee_user_id or "") == (asset.owner_user_id or "")
            and (open_checkout.location or "") == (asset.location or "")
        )
        due_date = AssetService.parse_optional_datetime(borrow_due_date)
        if same_assignment:
            open_checkout.dept_id = asset.dept_id
            open_checkout.assignee_name = AssetService.user_label(current_user, asset.owner_user_id) if asset.owner_user_id else None
            open_checkout.due_date = due_date
            open_checkout.remark = AssetService.join_notes(open_checkout.remark, remark)
            return

        if open_checkout:
            checkin_note = "重新领用"
            if previous_owner_user_id and previous_owner_user_id != asset.owner_user_id:
                checkin_note = f"转交给 {AssetService.user_label(current_user, asset.owner_user_id)}"
            AssetService.close_checkout(open_checkout, operator, asset.location, checkin_note)

        db.add(AssetCheckout(
            asset_id=asset.asset_id,
            checkout_type=to_status,
            assignee_user_id=AssetService.normalize_blank(asset.owner_user_id),
            assignee_name=AssetService.user_label(current_user, asset.owner_user_id) if asset.owner_user_id else None,
            dept_id=asset.dept_id,
            location=asset.location,
            due_date=due_date,
            checked_out_by=operator,
            remark=remark,
        ))

    @staticmethod
    def open_checkout_for_asset(db: Session, asset_id: str) -> AssetCheckout | None:
        return (
            db.query(AssetCheckout)
            .filter(AssetCheckout.asset_id == asset_id, AssetCheckout.status == "open")
            .order_by(AssetCheckout.checked_out_at.desc(), AssetCheckout.id.desc())
            .first()
        )

    @staticmethod
    def close_open_checkout(db: Session, asset: Asset, operator: str, location: str | None = None, remark: str | None = None) -> None:
        checkouts = (
            db.query(AssetCheckout)
            .filter(AssetCheckout.asset_id == asset.asset_id, AssetCheckout.status == "open")
            .with_for_update()
            .all()
        )
        for checkout in checkouts:
            AssetService.close_checkout(checkout, operator, location, remark)

    @staticmethod
    def close_checkout(checkout: AssetCheckout, operator: str, location: str | None = None, remark: str | None = None) -> None:
        checkout.status = "closed"
        checkout.open_token = None
        checkout.checked_in_at = utc_now()
        checkout.checked_in_by = operator
        checkout.checkin_location = AssetService.normalize_blank(location)
        checkout.checkin_remark = AssetService.join_notes(checkout.checkin_remark, remark)

    @staticmethod
    def list_checkouts(db: Session, asset_id: str, limit: int = 200, user_context: dict | None = None) -> list[dict]:
        AssetService.get_scoped_asset(db, asset_id, user_context)
        rows = (
            db.query(AssetCheckout)
            .filter(AssetCheckout.asset_id == asset_id)
            .order_by(AssetCheckout.checked_out_at.desc(), AssetCheckout.id.desc())
            .limit(min(max(limit, 1), 500))
            .all()
        )
        return [AssetService.checkout_out(row) for row in rows]

    @staticmethod
    def list_checkout_records(
        db: Session,
        page: int = 1,
        page_size: int = 20,
        keyword: str | None = None,
        status: str | None = None,
        checkout_type: str | None = None,
        assignee_user_id: str | None = None,
        dept_id: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        due_from: str | None = None,
        due_to: str | None = None,
        due_days: int = 7,
        user_context: dict | None = None,
    ) -> dict:
        now = utc_now()
        due_limit = now + AssetService.timedelta_days(max(due_days, 0))
        query = db.query(AssetCheckout, Asset).outerjoin(Asset, Asset.asset_id == AssetCheckout.asset_id)
        query = AssetService.apply_data_scope(query, user_context)
        clean_keyword = AssetService.normalize_blank(keyword)
        if clean_keyword:
            pattern = f"%{clean_keyword}%"
            query = query.filter(
                or_(
                    AssetCheckout.asset_id.like(pattern),
                    AssetCheckout.assignee_user_id.like(pattern),
                    AssetCheckout.assignee_name.like(pattern),
                    AssetCheckout.dept_id.like(pattern),
                    AssetCheckout.location.like(pattern),
                    AssetCheckout.remark.like(pattern),
                    Asset.asset_no.like(pattern),
                    Asset.name.like(pattern),
                    Asset.category.like(pattern),
                    Asset.brand.like(pattern),
                    Asset.model.like(pattern),
                    Asset.sn.like(pattern),
                )
            )
        if checkout_type:
            query = query.filter(AssetCheckout.checkout_type == checkout_type)
        if assignee_user_id:
            query = query.filter(AssetCheckout.assignee_user_id == assignee_user_id)
        if dept_id:
            query = query.filter(AssetCheckout.dept_id == dept_id)

        checked_out_from = AssetService.parse_optional_datetime(date_from)
        checked_out_to = AssetService.end_of_day(AssetService.parse_optional_datetime(date_to))
        due_start = AssetService.parse_optional_datetime(due_from)
        due_end = AssetService.end_of_day(AssetService.parse_optional_datetime(due_to))
        if checked_out_from:
            query = query.filter(AssetCheckout.checked_out_at >= checked_out_from)
        if checked_out_to:
            query = query.filter(AssetCheckout.checked_out_at <= checked_out_to)
        if due_start:
            query = query.filter(AssetCheckout.due_date >= due_start)
        if due_end:
            query = query.filter(AssetCheckout.due_date <= due_end)

        if status == "open":
            query = query.filter(AssetCheckout.status == "open")
        elif status == "closed":
            query = query.filter(AssetCheckout.status == "closed")
        elif status == "overdue":
            query = query.filter(AssetCheckout.status == "open", AssetCheckout.due_date.isnot(None), AssetCheckout.due_date < now)
        elif status == "due_soon":
            query = query.filter(AssetCheckout.status == "open", AssetCheckout.due_date.isnot(None), AssetCheckout.due_date >= now, AssetCheckout.due_date <= due_limit)
        elif status == "current":
            query = query.filter(AssetCheckout.status == "open")

        total = query.count()
        page = max(page, 1)
        page_size = min(max(page_size or 20, 1), 200)
        rows = (
            query.order_by(AssetCheckout.checked_out_at.desc(), AssetCheckout.id.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        base = db.query(AssetCheckout).join(Asset, Asset.asset_id == AssetCheckout.asset_id)
        base = AssetService.apply_data_scope(base, user_context)
        if checkout_type:
            base = base.filter(AssetCheckout.checkout_type == checkout_type)
        summary = {
            "open": base.filter(AssetCheckout.status == "open").count(),
            "closed": base.filter(AssetCheckout.status == "closed").count(),
            "overdue": base.filter(AssetCheckout.status == "open", AssetCheckout.due_date.isnot(None), AssetCheckout.due_date < now).count(),
            "due_soon": base.filter(AssetCheckout.status == "open", AssetCheckout.due_date.isnot(None), AssetCheckout.due_date >= now, AssetCheckout.due_date <= due_limit).count(),
        }
        return {
            "list": [AssetService.checkout_out(checkout, asset) for checkout, asset in rows],
            "total": total,
            "page": page,
            "page_size": page_size,
            "summary": summary,
        }

    @staticmethod
    def batch_checkout_assets(db: Session, payload: AssetBatchCheckoutCreate, operator: str = "system", user_context: dict | None = None) -> dict:
        return AssetService.batch_checkout_action(db, payload.asset_ids, "checkout", payload, operator, user_context)

    @staticmethod
    def batch_checkin_assets(db: Session, payload: AssetBatchCheckinCreate, operator: str = "system", user_context: dict | None = None) -> dict:
        return AssetService.batch_checkout_action(db, payload.asset_ids, "checkin", payload, operator, user_context)

    @staticmethod
    def batch_checkout_action(db: Session, asset_ids: list[str], action: str, payload: AssetCheckoutCreate | AssetCheckinCreate, operator: str, user_context: dict | None = None) -> dict:
        rows: list[dict] = []
        errors: list[dict] = []
        clean_ids = [AssetService.normalize_blank(asset_id) for asset_id in asset_ids if AssetService.normalize_blank(asset_id)]
        for asset_id in dict.fromkeys(clean_ids):
            try:
                if action == "checkout":
                    asset = AssetService.checkout_asset(db, asset_id, payload, operator, user_context)
                else:
                    asset = AssetService.checkin_asset(db, asset_id, payload, operator, user_context)
                rows.append(asset)
            except (AssetValidationError, ValueError) as exc:
                # 失败条目在校验前可能已修改 ORM 对象，必须回滚，否则残留变更会随下一条的 commit 一起提交
                db.rollback()
                errors.append({"asset_id": asset_id, "message": str(exc)})
        return {"success": len(rows), "failed": len(errors), "assets": rows, "errors": errors}

    @staticmethod
    def checkout_out(row: AssetCheckout, asset: Asset | None = None) -> dict:
        now = utc_now()
        due_date = row.due_date
        is_open = row.status == "open"
        is_overdue = bool(is_open and due_date and due_date < now)
        return {
            "id": row.id,
            "asset_id": row.asset_id,
            "asset_no": asset.asset_no if asset else None,
            "asset_name": asset.name if asset else None,
            "asset_category": asset.category if asset else None,
            "asset_status": asset.status if asset else None,
            "checkout_type": row.checkout_type,
            "assignee_user_id": row.assignee_user_id,
            "assignee_name": row.assignee_name,
            "dept_id": row.dept_id,
            "location": row.location,
            "due_date": row.due_date,
            "status": row.status,
            "is_overdue": is_overdue,
            "days_overdue": (now.date() - due_date.date()).days if is_overdue and due_date else 0,
            "checked_out_at": row.checked_out_at,
            "checked_out_by": row.checked_out_by,
            "checked_in_at": row.checked_in_at,
            "checked_in_by": row.checked_in_by,
            "checkin_location": row.checkin_location,
            "remark": row.remark,
            "checkin_remark": row.checkin_remark,
        }

    @staticmethod
    def parse_optional_datetime(value: str | None) -> datetime | None:
        clean = AssetService.normalize_blank(value)
        if not clean:
            return None
        try:
            return app_datetime_to_utc(datetime.fromisoformat(clean.replace("Z", "+00:00")))
        except ValueError:
            try:
                return app_datetime_to_utc(datetime.strptime(clean, "%Y-%m-%d"))
            except ValueError as exc:
                raise AssetValidationError(f"日期格式不正确：{clean}，请使用 YYYY-MM-DD") from exc

    @staticmethod
    def end_of_day(value: datetime | None) -> datetime | None:
        if not value:
            return None
        return app_end_of_day(value)

    @staticmethod
    def timedelta_days(days: int):
        from datetime import timedelta

        return timedelta(days=days)

    @staticmethod
    def join_notes(original: str | None, note: str | None) -> str | None:
        clean_original = AssetService.normalize_blank(original)
        clean_note = AssetService.normalize_blank(note)
        if not clean_note:
            return clean_original
        if not clean_original:
            return clean_note
        if clean_note in clean_original:
            return clean_original
        return f"{clean_original}；{clean_note}"

    @staticmethod
    def notify_status_change(
        db: Session,
        asset: Asset,
        from_status: str | None,
        to_status: str,
        operator: str,
        previous_user: UserDirectory | None = None,
        previous_owner_user_id: str | None = None,
        current_user: UserDirectory | None = None,
    ) -> None:
        now_text = format_app_datetime()
        if to_status == "in_stock":
            NotificationService.send_event(
                db,
                "inbound",
                "资产入库完成",
                [
                    f"资产名称：{AssetService.asset_display_name(asset)}",
                    *NotificationService.asset_identity_lines(asset),
                    f"状态变更：{AssetService.status_label(from_status)} -> {AssetService.status_label(to_status)}",
                    f"入库位置：{asset.location or '-'}",
                    f"原责任人：{AssetService.user_label(previous_user, previous_owner_user_id)}",
                    f"操作人：{operator}",
                    f"操作时间：{now_text}",
                ],
            )
        elif to_status in {"in_use", "borrowed", "out_stock"}:
            owner_label = AssetService.user_label(current_user, asset.owner_user_id)
            owner_key = "公用设备位置" if to_status == "out_stock" and not AssetService.normalize_blank(asset.owner_user_id) else AssetService.outbound_owner_key(to_status)
            NotificationService.send_event(
                db,
                "outbound",
                "资产出库完成",
                [
                    f"资产名称：{AssetService.asset_display_name(asset)}",
                    *NotificationService.asset_identity_lines(asset),
                    f"状态变更：{AssetService.status_label(from_status)} -> {AssetService.status_label(to_status)}",
                    f"{owner_key}：{asset.location if owner_key == '公用设备位置' else owner_label}",
                    f"所属部门：{current_user.dept_name or current_user.dept_id if current_user else asset.dept_id or '-'}",
                    f"使用位置：{asset.location or '-'}",
                    f"操作人：{operator}",
                    f"操作时间：{now_text}",
                ],
            )

    @staticmethod
    def asset_display_name(asset: Asset) -> str:
        parts = [asset.name, asset.brand, asset.model]
        return " / ".join([str(part).strip() for part in parts if part]) or asset.asset_id

    @staticmethod
    def status_label(value: str | None) -> str:
        return {
            "pending_acceptance": "待验收",
            "in_stock": "在库",
            "in_use": "在用",
            "idle": "闲置",
            "borrowed": "借出",
            "repair": "维修中",
            "out_stock": "已出库",
            "ready_scrap": "待报废",
            "pending_scrap": "待处置登记",
            "scrapped": "已报废",
            "disposed": "已处置",
            "lost": "已丢失",
        }.get(value or "", value or "-")

    @staticmethod
    def outbound_owner_key(to_status: str) -> str:
        return {
            "in_use": "领用人",
            "borrowed": "借用人",
            "out_stock": "出库责任人",
        }.get(to_status, "责任人")

    @staticmethod
    def user_label(user: UserDirectory | None, fallback: str | None) -> str:
        if user:
            return user.display_name or user.username or user.user_id
        return AssetService.normalize_blank(fallback) or "-"

    @staticmethod
    def inventory_lifecycle_remark(
        to_status: str,
        remark: str | None,
        previous_owner_user_id: str | None,
        previous_user: UserDirectory | None,
        owner_user_id: str | None,
        owner_user: UserDirectory | None,
        location: str | None = None,
        borrow_due_date: str | None = None,
    ) -> dict | str | None:
        base = AssetService.normalize_blank(remark)
        labels = {
            "in_stock": ("退回人", AssetService.user_label(previous_user, previous_owner_user_id)),
            "in_use": ("领用人", AssetService.user_label(owner_user, owner_user_id)),
            "borrowed": ("借用人", AssetService.user_label(owner_user, owner_user_id)),
            "out_stock": ("出库责任人", AssetService.user_label(owner_user, owner_user_id)),
        }
        if to_status not in labels:
            return base or None
        key, value = labels[to_status]
        if to_status == "in_stock":
            key, value = "入库地址", AssetService.normalize_blank(location) or "-"
        if to_status == "out_stock" and not AssetService.normalize_blank(owner_user_id):
            key, value = "公用设备位置", AssetService.normalize_blank(location) or "-"
        details = []
        if base:
            details.append(f"操作原因: {base}")
        details.append(f"{key}: {value}")
        if to_status != "in_stock":
            details.append(f"原责任人: {AssetService.user_label(previous_user, previous_owner_user_id)}")
            details.append(f"新责任人: {AssetService.user_label(owner_user, owner_user_id)}")
        details.append(f"位置: {AssetService.normalize_blank(location) or '-'}")
        if to_status == "borrowed":
            due_date = AssetService.normalize_blank(borrow_due_date)
            if due_date:
                details.append(f"借用到期时间: {due_date}")
        operation_object = {
            "in_stock": "资产归还入库",
            "in_use": "资产领用",
            "borrowed": "资产借用",
            "out_stock": "资产出库",
        }.get(to_status, "资产状态流转")
        return LifecycleService.structured_remark(
            reason=base or operation_object,
            object=operation_object,
            previous_owner=AssetService.user_label(previous_user, previous_owner_user_id),
            new_owner=AssetService.user_label(owner_user, owner_user_id),
            location=AssetService.normalize_blank(location) or "-",
            due_date=AssetService.normalize_blank(borrow_due_date) if to_status == "borrowed" else "",
            extra={
                "owner_label": key,
                "owner_value": value,
                "status": to_status,
            },
        )

    @staticmethod
    def find_user(db: Session, value: str | None, active_only: bool = False) -> UserDirectory | None:
        if not value:
            return None
        clean_value = str(value).strip()
        candidates = [clean_value]
        if clean_value.startswith("ldap:"):
            candidates.append(clean_value.removeprefix("ldap:"))
        lowered = clean_value.lower()
        if "cn=" in lowered:
            cn_part = lowered.split("cn=", 1)[1].split(",", 1)[0]
            if cn_part:
                candidates.append(cn_part)
        candidate_keys = {item.casefold() for item in candidates if item}
        exact_query = (
            db.query(UserDirectory)
            .filter(
                or_(
                    func.lower(UserDirectory.user_id).in_(candidate_keys),
                    func.lower(UserDirectory.username).in_(candidate_keys),
                    func.lower(UserDirectory.external_id).in_(candidate_keys),
                    func.lower(UserDirectory.email).in_(candidate_keys),
                    func.lower(UserDirectory.display_name).in_(candidate_keys),
                )
            )
        )
        if active_only:
            exact_query = exact_query.filter(func.lower(UserDirectory.status) == "active")
        exact = exact_query.first()
        if exact:
            return exact
        compound_candidates = [
            item.strip()
            for item in re.split(r"[\s\-_/\\|,;:，；：()（）]+", clean_value)
            if item.strip()
        ]
        if not compound_candidates:
            return None
        compound_candidate_keys = {item.casefold() for item in compound_candidates}
        compound_query = (
            db.query(UserDirectory)
            .filter(
                or_(
                    func.lower(UserDirectory.user_id).in_(compound_candidate_keys),
                    func.lower(UserDirectory.username).in_(compound_candidate_keys),
                    func.lower(UserDirectory.external_id).in_(compound_candidate_keys),
                    func.lower(UserDirectory.email).in_(compound_candidate_keys),
                    func.lower(UserDirectory.display_name).in_(compound_candidate_keys),
                )
            )
        )
        if active_only:
            compound_query = compound_query.filter(func.lower(UserDirectory.status) == "active")
        return compound_query.first()

    @staticmethod
    def sync_owner_department(db: Session, asset: Asset) -> UserDirectory | None:
        user = AssetService.find_user(db, asset.owner_user_id)
        if not user:
            return None
        asset.owner_user_id = user.user_id
        asset.dept_id = user.dept_id or user.dept_name or asset.dept_id
        return user

    @staticmethod
    def users_by_identity(db: Session) -> dict[str, UserDirectory]:
        users = db.query(UserDirectory).all()
        mapping: dict[str, UserDirectory] = {}
        for user in users:
            for value in [user.user_id, user.username, user.external_id, user.email, user.display_name]:
                if value:
                    mapping[value] = user
        return mapping

    @staticmethod
    def to_out(asset: Asset, user: UserDirectory | None = None, db: Session | None = None, residual_config: dict | None = None) -> dict:
        user = user or None
        return {
            "asset_id": asset.asset_id,
            "asset_no": asset.asset_no,
            "company": asset.company or AssetService.DEFAULT_COMPANY,
            "name": asset.name,
            "category": asset.category,
            "brand": asset.brand,
            "model": asset.model,
            "sn": asset.sn,
            "config": asset.config,
            "purchase_price": asset.purchase_price,
            "current_residual_value": AssetResidualService.calculate_asset(asset, db=db, residual_config=residual_config),
            "purchase_date": asset.purchase_date,
            "purchase_approval_no": asset.purchase_approval_no,
            "purchase_supplier_name": asset.purchase_supplier_name,
            "warranty_expire_date": asset.warranty_expire_date,
            "warranty_months": asset.warranty_months,
            "status": asset.status,
            "owner_user_id": asset.owner_user_id,
            "owner_display_name": user.display_name if user else None,
            "owner_username": user.username if user else None,
            "dept_id": asset.dept_id,
            "dept_name": user.dept_name if user else None,
            "location": asset.location,
            "remark": asset.remark,
            "created_at": asset.created_at,
        }

    @staticmethod
    def parse_datetime(value: Any) -> datetime | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return app_datetime_to_utc(value)
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
            try:
                return app_datetime_to_utc(datetime.strptime(text, fmt))
            except ValueError:
                continue
        try:
            return app_datetime_to_utc(datetime.fromisoformat(text))
        except ValueError:
            return None

    @staticmethod
    def parse_float(value: Any, field_name: str = "数字") -> float:
        if value in (None, ""):
            return 0
        if isinstance(value, int | float):
            return float(value)
        text = str(value).strip()
        if not text:
            return 0
        normalized = text.replace(",", "").replace("￥", "").replace("¥", "").replace("元", "").replace(" ", "")
        try:
            return float(normalized)
        except ValueError as exc:
            raise AssetValidationError(f"{field_name}格式不正确：{text}，请填写数字，例如 1380.00") from exc

    @staticmethod
    def parse_int(value: Any) -> int | None:
        if value in (None, ""):
            return None
        try:
            return int(AssetService.parse_float(value))
        except (TypeError, ValueError):
            return None

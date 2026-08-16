import csv
from datetime import datetime, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import FileResponse, PlainTextResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.time import app_now, format_app_datetime, utc_now
from app.core.config import get_settings
from app.core.database import get_db
from app.core.security import can_view_all_data, scoped_dept_id, scoped_user_identities, operator_from_request, user_context_from_request
from app.models.asset import Asset
from app.models.checkout import AssetCheckout
from app.models.lifecycle import Lifecycle
from app.models.repair import RepairRecord
from app.models.report import AuditReportArchive
from app.models.scrap import ScrapRequest
from app.models.stocktake import StocktakeItem, StocktakeTask
from app.models.user import UserDirectory
from app.services.asset_service import AssetService
from app.services.asset_residual_service import AssetResidualService
from app.services.number_service import NumberService
from app.services.audit_engine import AuditEngine
from app.reports.generator import AuditReportGenerator


router = APIRouter(prefix="/reports", tags=["Reports"])


def audit_report_out(row: AuditReportArchive) -> dict:
    return {
        "id": row.report_no,
        "archive_id": row.id,
        "name": row.name,
        "type": "审计报告",
        "status": row.status,
        "created_at": date_text(row.created_at),
        "total_assets": row.total_assets,
        "risk_score": row.risk_score,
        "violation_count": row.violation_count,
        "created_by": row.created_by or "",
        "has_pdf": bool(row.pdf_path),
        "has_xlsx": bool(row.xlsx_path),
    }


@router.get("/audit-reports")
def list_audit_report_archives(request: Request, page: int = 1, page_size: int = 50, db: Session = Depends(get_db)):
    clean_page = max(page, 1)
    clean_page_size = min(max(page_size, 1), 200)
    query = scoped_audit_archives_query(db, request).order_by(AuditReportArchive.created_at.desc(), AuditReportArchive.id.desc())
    total = query.count()
    rows = query.offset((clean_page - 1) * clean_page_size).limit(clean_page_size).all()
    return {"list": [audit_report_out(row) for row in rows], "total": total, "page": clean_page, "page_size": clean_page_size}


@router.post("/audit-reports")
def create_audit_report_archive(request: Request, db: Session = Depends(get_db)):
    asset_ids = set(scoped_asset_id_query(db, request))
    result = AuditEngine(db).run(asset_ids=asset_ids)
    year = app_now().year
    report_no = NumberService.next(db, f"audit_report:{year}", f"AR-{year}-", 6)
    output_dir = Path(get_settings().upload_dir) / "reports" / "audit-archives"
    output_dir.mkdir(parents=True, exist_ok=True)

    generated_html = Path(AuditReportGenerator().generate(result))
    html_path = output_dir / f"{report_no}.html"
    html_path.write_text(generated_html.read_text(encoding="utf-8"), encoding="utf-8")

    from app.api.audit import generate_audit_pdf

    pdf_source = Path(generate_audit_pdf(result))
    pdf_path = output_dir / f"{report_no}.pdf"
    pdf_path.write_bytes(pdf_source.read_bytes())

    xlsx_path = output_dir / f"{report_no}.xlsx"
    build_audit_workbook(result).save(xlsx_path)

    row = AuditReportArchive(
        report_no=report_no,
        name=f"审计报告 {format_app_datetime(fmt='%Y-%m-%d %H:%M')}",
        report_type="audit",
        status="已生成",
        total_assets=int(result.get("total_assets") or 0),
        risk_score=int(result.get("risk_score") or 0),
        violation_count=len(result.get("violations") or []),
        html_path=str(html_path),
        pdf_path=str(pdf_path),
        xlsx_path=str(xlsx_path),
        created_by=operator_from_request(request),
        scope_key=report_scope_key(user_context_from_request(request)),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {**audit_report_out(row), "html": html_path.read_text(encoding="utf-8")}


@router.get("/audit-reports/{report_no}/html")
def get_archived_audit_report_html(report_no: str, request: Request, db: Session = Depends(get_db)):
    row = get_audit_report_archive(db, report_no, request)
    path = Path(row.html_path)
    if not path.exists():
        raise HTTPException(status_code=404, detail="report file not found")
    return FileResponse(path, media_type="text/html; charset=utf-8", filename=f"{report_no}.html")


@router.get("/audit-reports/{report_no}/pdf")
def download_archived_audit_report_pdf(report_no: str, request: Request, db: Session = Depends(get_db)):
    row = get_audit_report_archive(db, report_no, request)
    path = Path(row.pdf_path or "")
    if not path.exists():
        raise HTTPException(status_code=404, detail="report pdf not found")
    return FileResponse(path, media_type="application/pdf", filename=f"{report_no}.pdf")


@router.get("/audit-reports/{report_no}/xlsx")
def download_archived_audit_report_xlsx(report_no: str, request: Request, db: Session = Depends(get_db)):
    row = get_audit_report_archive(db, report_no, request)
    path = Path(row.xlsx_path or "")
    if not path.exists():
        raise HTTPException(status_code=404, detail="report excel not found")
    return FileResponse(path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"{report_no}.xlsx")


def get_audit_report_archive(db: Session, report_no: str, request: Request) -> AuditReportArchive:
    row = scoped_audit_archives_query(db, request).filter(AuditReportArchive.report_no == report_no).first()
    if not row:
        raise HTTPException(status_code=404, detail="report not found")
    return row


@router.get("/assets.csv")
def export_assets_csv(request: Request, db: Session = Depends(get_db)):
    rows = [
        [asset.asset_id, asset.asset_no, asset_info(asset), asset.name, asset.category, asset.brand, asset.model, asset.sn, asset.status, asset.owner_user_id, asset.dept_id, asset.location, asset.purchase_price, AssetResidualService.calculate_asset(asset, db=db), asset.purchase_approval_no, asset.purchase_supplier_name]
        for asset in scoped_assets_query(db, request).order_by(Asset.asset_id.asc()).all()
    ]
    return csv_response("assets.csv", ["资产ID", "资产编号", "资产信息", "名称", "类型", "品牌", "型号", "SN", "状态", "责任人", "部门", "位置", "采购价值", "当前残值", "采购审批单号", "采购供应商"], rows)


@router.get("/department-assets.csv")
def export_department_assets_csv(request: Request, db: Session = Depends(get_db)):
    rows = [
        [
            asset.dept_id or "",
            asset.asset_id,
            asset.asset_no or "",
            asset_info(asset),
            asset.name,
            asset.category,
            asset.brand or "",
            asset.model or "",
            asset.sn or "",
            asset.status or "",
            asset.owner_user_id or "",
            asset.location or "",
            money(asset.purchase_price),
            money(AssetResidualService.calculate_asset(asset, db=db)),
            date_text(asset.created_at),
        ]
        for asset in scoped_assets_query(db, request).order_by(Asset.dept_id.asc(), Asset.asset_id.asc()).all()
    ]
    return csv_response("department-assets.csv", ["部门", "资产ID", "资产编号", "资产信息", "名称", "类型", "品牌", "型号", "SN", "状态", "责任人", "位置", "资产价值", "当前残值", "创建时间"], rows)


@router.get("/person-holdings.csv")
def export_person_holdings_csv(request: Request, db: Session = Depends(get_db)):
    users = users_by_identity(db)
    rows = []
    query = (
        scoped_assets_query(db, request)
        .filter(Asset.owner_user_id.isnot(None), Asset.owner_user_id != "", ~Asset.status.in_(["scrapped", "disposed", "lost"]))
        .order_by(Asset.owner_user_id.asc(), Asset.asset_id.asc())
    )
    for asset in query.all():
        user = users.get(asset.owner_user_id or "")
        rows.append([
            asset.owner_user_id or "",
            user.display_name if user else "",
            (user.dept_name or user.dept_id) if user else asset.dept_id or "",
            asset.asset_id,
            asset.asset_no or "",
            asset_info(asset),
            asset.name,
            asset.category,
            asset.status or "",
            asset.location or "",
            money(asset.purchase_price),
            money(AssetResidualService.calculate_asset(asset, db=db)),
            date_text(asset.created_at),
        ])
    return csv_response("person-holdings.csv", ["人员ID", "姓名", "部门", "资产ID", "资产编号", "资产信息", "名称", "类型", "状态", "位置", "资产价值", "当前残值", "创建时间"], rows)


@router.get("/overdue-borrowings.csv")
def export_overdue_borrowings_csv(request: Request, db: Session = Depends(get_db)):
    now = utc_now()
    scoped_asset_ids = scoped_asset_id_query(db, request)
    rows = []
    query = (
        db.query(AssetCheckout, Asset)
        .join(Asset, AssetCheckout.asset_id == Asset.asset_id)
        .filter(Asset.asset_id.in_(scoped_asset_ids), AssetCheckout.status == "open", AssetCheckout.due_date.isnot(None), AssetCheckout.due_date < now)
        .order_by(AssetCheckout.due_date.asc(), AssetCheckout.asset_id.asc())
    )
    for checkout, asset in query.all():
        rows.append([
            checkout.asset_id,
            asset.asset_no or "",
            asset_info(asset),
            asset.name,
            checkout.checkout_type,
            checkout.assignee_user_id or "",
            checkout.assignee_name or "",
            checkout.dept_id or asset.dept_id or "",
            date_text(checkout.checked_out_at),
            date_text(checkout.due_date),
            (now.date() - checkout.due_date.date()).days if checkout.due_date else 0,
            checkout.location or asset.location or "",
            checkout.remark or "",
        ])
    return csv_response("overdue-borrowings.csv", ["资产ID", "资产编号", "资产信息", "名称", "领用类型", "持有人ID", "持有人", "部门", "领用时间", "到期时间", "逾期天数", "位置", "备注"], rows)


@router.get("/borrowings.csv")
def export_borrowings_csv(request: Request, db: Session = Depends(get_db)):
    scoped_asset_ids = scoped_asset_id_query(db, request)
    rows = [[
        checkout.asset_id,
        asset.asset_no or "",
        asset_info(asset),
        checkout.checkout_type,
        checkout.assignee_user_id or "",
        checkout.assignee_name or "",
        checkout.dept_id or asset.dept_id or "",
        checkout.status,
        date_text(checkout.checked_out_at),
        date_text(checkout.due_date),
        date_text(checkout.checked_in_at),
        checkout.location or asset.location or "",
        checkout.checkin_location or "",
        checkout.remark or "",
        checkout.checkin_remark or "",
    ] for checkout, asset in (
        db.query(AssetCheckout, Asset)
        .join(Asset, AssetCheckout.asset_id == Asset.asset_id)
        .filter(Asset.asset_id.in_(scoped_asset_ids))
        .order_by(AssetCheckout.checked_out_at.desc(), AssetCheckout.id.desc())
        .all()
    )]
    return csv_response("borrowings.csv", ["资产ID", "资产编号", "资产信息", "借用类型", "人员ID", "人员", "部门", "记录状态", "借用时间", "计划归还", "实际归还", "使用位置", "归还位置", "借用备注", "归还备注"], rows)


@router.get("/repairs.csv")
def export_repairs_csv(request: Request, db: Session = Depends(get_db)):
    scoped_asset_ids = scoped_asset_id_query(db, request)
    rows = [[
        repair.repair_no,
        repair.asset_id,
        asset.asset_no or "",
        asset_info(asset),
        repair.repair_type,
        repair.fault_reason,
        money(repair.repair_cost),
        repair.vendor or "",
        repair.status,
        repair.repair_result or "",
        date_text(repair.repair_time),
        date_text(repair.finish_time),
        repair.operator or "",
        repair.remark or "",
    ] for repair, asset in (
        db.query(RepairRecord, Asset)
        .join(Asset, RepairRecord.asset_id == Asset.asset_id)
        .filter(Asset.asset_id.in_(scoped_asset_ids))
        .order_by(RepairRecord.repair_time.desc(), RepairRecord.id.desc())
        .all()
    )]
    return csv_response("repairs.csv", ["维修单号", "资产ID", "资产编号", "资产信息", "维修类型", "故障原因", "维修费用", "维修供应商", "状态", "维修结果", "送修时间", "完成时间", "登记人", "备注"], rows)


@router.get("/stocktake-items.csv")
def export_stocktake_items_csv(request: Request, db: Session = Depends(get_db)):
    scoped_asset_ids = scoped_asset_id_query(db, request)
    rows = [[
        task.id,
        task.name,
        task.status,
        item.asset_id,
        asset.asset_no or "",
        asset_info(asset),
        item.sn or asset.sn or "",
        item.book_status or "",
        item.book_owner_user_id or "",
        item.actual_owner_user_id or "",
        item.book_location or "",
        item.actual_location or "",
        item.result,
        item.checker or "",
        date_text(item.checked_at),
        item.review_status,
        item.reviewed_by or "",
        date_text(item.reviewed_at),
        item.remark or "",
    ] for item, task, asset in (
        db.query(StocktakeItem, StocktakeTask, Asset)
        .join(StocktakeTask, StocktakeItem.task_id == StocktakeTask.id)
        .join(Asset, StocktakeItem.asset_id == Asset.asset_id)
        .filter(Asset.asset_id.in_(scoped_asset_ids))
        .order_by(StocktakeTask.created_at.desc(), StocktakeItem.id.asc())
        .all()
    )]
    return csv_response("stocktake-items.csv", ["盘点任务ID", "盘点任务", "任务状态", "资产ID", "资产编号", "资产信息", "SN", "账面状态", "账面使用人", "实际使用人", "账面位置", "实际位置", "盘点结果", "盘点人", "盘点时间", "复核状态", "复核人", "复核时间", "备注"], rows)


@router.get("/purchase-assets.csv")
def export_purchase_assets_csv(request: Request, db: Session = Depends(get_db)):
    query = (
        scoped_assets_query(db, request)
        .filter(Asset.purchase_approval_no.isnot(None), Asset.purchase_approval_no != "")
        .order_by(Asset.purchase_approval_no.asc(), Asset.asset_id.asc())
    )
    rows = [[
        (asset.config or {}).get("purchase_no") or "",
        asset.purchase_approval_no or "",
        asset.asset_id,
        asset.asset_no or "",
        asset_info(asset),
        asset.sn or "",
        asset.category,
        asset.status,
        asset.purchase_supplier_name or "",
        money(asset.purchase_price),
        date_text(asset.purchase_date),
        asset.owner_user_id or "",
        asset.dept_id or "",
        asset.location or "",
    ] for asset in query.all()]
    return csv_response("purchase-assets.csv", ["采购单号", "采购审批单号", "资产ID", "资产编号", "资产信息", "SN", "类型", "状态", "采购供应商", "采购价值", "采购日期", "责任人", "部门", "位置"], rows)


@router.get("/lifecycle.csv")
def export_lifecycle_csv(request: Request, db: Session = Depends(get_db)):
    scoped_asset_ids = scoped_asset_id_query(db, request)
    rows = [[
        lifecycle.id,
        lifecycle.asset_id,
        asset.asset_no or "",
        asset_info(asset),
        lifecycle.action_type,
        lifecycle.from_status or "",
        lifecycle.to_status or "",
        lifecycle.operator or "",
        lifecycle.remark or "",
        date_text(lifecycle.timestamp),
    ] for lifecycle, asset in (
        db.query(Lifecycle, Asset)
        .join(Asset, Lifecycle.asset_id == Asset.asset_id)
        .filter(Asset.asset_id.in_(scoped_asset_ids))
        .order_by(Lifecycle.timestamp.desc(), Lifecycle.id.desc())
        .all()
    )]
    return csv_response("asset-lifecycle.csv", ["流水ID", "资产ID", "资产编号", "资产信息", "操作类型", "原状态", "新状态", "操作人", "说明", "操作时间"], rows)


@router.get("/warranty-expiring.csv")
def export_warranty_expiring_csv(request: Request, days: int = 90, db: Session = Depends(get_db)):
    now = utc_now()
    end_at = now + timedelta(days=max(days, 0))
    rows = [
        [
            asset.asset_id,
            asset.asset_no or "",
            asset_info(asset),
            asset.name,
            asset.category,
            asset.brand or "",
            asset.model or "",
            asset.sn or "",
            asset.status or "",
            asset.owner_user_id or "",
            asset.dept_id or "",
            asset.location or "",
            date_text(asset.purchase_date),
            date_text(asset.warranty_expire_date),
            (asset.warranty_expire_date.date() - now.date()).days if asset.warranty_expire_date else "",
        ]
        for asset in scoped_assets_query(db, request)
        .filter(Asset.warranty_expire_date.isnot(None), Asset.warranty_expire_date >= now, Asset.warranty_expire_date <= end_at)
        .order_by(Asset.warranty_expire_date.asc(), Asset.asset_id.asc())
        .all()
    ]
    return csv_response("warranty-expiring.csv", ["资产ID", "资产编号", "资产信息", "名称", "类型", "品牌", "型号", "SN", "状态", "责任人", "部门", "位置", "采购日期", "质保到期", "剩余天数"], rows)


@router.get("/scrap-disposal-ledger.csv")
def export_scrap_disposal_ledger_csv(request: Request, db: Session = Depends(get_db)):
    scoped_asset_ids = scoped_asset_id_query(db, request)
    rows = []
    query = (
        db.query(ScrapRequest, Asset)
        .join(Asset, ScrapRequest.asset_id == Asset.asset_id)
        .filter(ScrapRequest.asset_id.in_(scoped_asset_ids))
        .order_by(ScrapRequest.created_at.desc(), ScrapRequest.id.desc())
    )
    for row, asset in query.all():
        rows.append([
            row.request_no,
            row.asset_id,
            asset.asset_no or "",
            asset_info(asset),
            row.asset_name,
            row.asset_sn or "",
            row.category or "",
            row.brand or "",
            row.model or "",
            row.owner_user_id or "",
            row.dept_id or "",
            row.location or "",
            money(row.purchase_price),
            row.status,
            row.applicant or "",
            row.reason or "",
            date_text(row.retirement_date),
            row.retirement_approval_no or "",
            row.disposal_method or "",
            row.dispose_recipient_name or row.dispose_recipient_user_id or "",
            money(row.estimated_residual_value),
            money(row.final_residual_value),
            row.disposed_by or "",
            date_text(row.disposed_at),
            row.disposal_remark or "",
            date_text(row.created_at),
        ])
    return csv_response("scrap-disposal-ledger.csv", ["报废单号", "资产ID", "资产编号", "资产信息", "资产名称", "SN", "类型", "品牌", "型号", "责任人", "部门", "位置", "采购价值", "状态", "申请人", "报废原因", "退役时间", "退役审批单号", "处置方式", "报废领走人", "预计残值", "最终残值", "处置人", "处置时间", "处置备注", "申请时间"], rows)


@router.get("/audit-report.xlsx")
def export_audit_report_excel(request: Request, db: Session = Depends(get_db)):
    result = AuditEngine(db).run(asset_ids=set(scoped_asset_id_query(db, request)))
    workbook = Workbook()
    build_audit_summary_sheet(workbook.active, result)
    build_audit_violations_sheet(workbook.create_sheet("风险明细"), result.get("violations") or [])
    build_audit_responses_sheet(workbook.create_sheet("审计答复"), result.get("responses") or [])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"audit-report-{app_now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/assets.pdf")
def export_assets_pdf(request: Request, db: Session = Depends(get_db)):
    output_dir = Path(get_settings().upload_dir) / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "assets.pdf"
    c = canvas.Canvas(str(output_path), pagesize=A4)
    width, height = A4
    y = height - 40
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "ITAM Asset Report")
    y -= 28
    c.setFont("Helvetica", 9)
    for asset in scoped_assets_query(db, request).order_by(Asset.asset_id.asc()).all():
        line = f"{asset.asset_id} | {asset.asset_no or '-'} | {asset_info(asset)} | {asset.category} | {asset.status} | {asset.owner_user_id or '-'} | {asset.location or '-'}"
        c.drawString(40, y, line[:120])
        y -= 16
        if y < 40:
            c.showPage()
            c.setFont("Helvetica", 9)
            y = height - 40
    c.save()
    return FileResponse(output_path, filename="assets.pdf", media_type="application/pdf")


@router.get("/analytics")
def report_analytics(request: Request, db: Session = Depends(get_db)):
    scoped_asset_ids = [row[0] for row in scoped_assets_query(db, request).with_entities(Asset.asset_id).all()]
    if not scoped_asset_ids:
        return {
            "department_occupancy": [],
            "per_capita_value": [],
            "idle_trend": [],
            "repair_cost_trend": [],
            "stocktake_diff_trend": [],
        }
    department_rows = (
        db.query(
            Asset.dept_id,
            func.count(Asset.asset_id),
            func.coalesce(func.sum(Asset.purchase_price), 0),
            func.count(func.distinct(Asset.owner_user_id)),
        )
        .filter(Asset.asset_id.in_(scoped_asset_ids), ~Asset.status.in_(["scrapped", "disposed", "lost"]))
        .group_by(Asset.dept_id)
        .order_by(func.count(Asset.asset_id).desc())
        .all()
    )
    department_occupancy = [
        {
            "dept_id": dept_id or "未绑定",
            "asset_count": int(asset_count or 0),
            "asset_value": float(asset_value or 0),
            "owner_count": int(owner_count or 0),
            "per_capita_value": float(asset_value or 0) / max(int(owner_count or 0), 1),
        }
        for dept_id, asset_count, asset_value, owner_count in department_rows
    ]
    months = month_windows(6)
    # 批量聚合替代每月一次查询（6 个月 × 3 个趋势 = 18 次 → 3 次）
    month_labels = [label for label, _start, _end in months]

    idle_rows = dict(
        db.query(
            func.date_format(Asset.created_at, "%Y-%m"),
            func.count(Asset.asset_id),
        )
        .filter(
            Asset.asset_id.in_(scoped_asset_ids),
            Asset.status.in_(["idle", "in_stock"]),
            Asset.created_at >= months[0][1],
            Asset.created_at <= months[-1][2],
        )
        .group_by(func.date_format(Asset.created_at, "%Y-%m"))
        .all()
    )
    # 窗口外（早于首月）创建、当前仍为 idle/in_stock 的资产作为基数，
    # 与逐月 count(created_at <= 月末) 的原口径保持一致。
    idle_base = int(
        db.query(func.count(Asset.asset_id))
        .filter(
            Asset.asset_id.in_(scoped_asset_ids),
            Asset.status.in_(["idle", "in_stock"]),
            Asset.created_at < months[0][1],
        )
        .scalar()
        or 0
    )
    idle_cumulative = idle_base
    idle_trend = []
    for month, _start_at, _end_at in months:
        idle_cumulative += int(idle_rows.get(month, 0) or 0)
        idle_trend.append({"month": month, "count": idle_cumulative})

    repair_cost_rows = dict(
        db.query(
            func.date_format(RepairRecord.repair_time, "%Y-%m"),
            func.coalesce(func.sum(RepairRecord.repair_cost), 0),
        )
        .filter(
            RepairRecord.asset_id.in_(scoped_asset_ids),
            RepairRecord.repair_time >= months[0][1],
            RepairRecord.repair_time < months[-1][2],
        )
        .group_by(func.date_format(RepairRecord.repair_time, "%Y-%m"))
        .all()
    )
    repair_cost_trend = [
        {"month": month, "cost": float(repair_cost_rows.get(month, 0) or 0)}
        for month, _start_at, _end_at in months
    ]

    stocktake_diff_rows = dict(
        db.query(
            func.date_format(StocktakeTask.created_at, "%Y-%m"),
            func.count(StocktakeItem.id),
        )
        .join(StocktakeTask, StocktakeItem.task_id == StocktakeTask.id)
        .filter(
            StocktakeItem.asset_id.in_(scoped_asset_ids),
            StocktakeTask.created_at >= months[0][1],
            StocktakeTask.created_at < months[-1][2],
            StocktakeItem.result != "正常",
        )
        .group_by(func.date_format(StocktakeTask.created_at, "%Y-%m"))
        .all()
    )
    stocktake_diff_trend = [
        {"month": month, "diff_count": int(stocktake_diff_rows.get(month, 0) or 0)}
        for month, _start_at, _end_at in months
    ]
    return {
        "department_occupancy": department_occupancy,
        "per_capita_value": department_occupancy,
        "idle_trend": idle_trend,
        "repair_cost_trend": repair_cost_trend,
        "stocktake_diff_trend": stocktake_diff_trend,
    }


def scoped_assets_query(db: Session, request: Request):
    return AssetService.apply_data_scope(db.query(Asset), user_context_from_request(request))


def report_scope_key(user_context: dict | None) -> str:
    if can_view_all_data(user_context):
        return "global"
    dept_id = scoped_dept_id(user_context)
    role = ((user_context or {}).get("role") or "").lower()
    if role in {"dept_manager", "department_manager", "manager"} and dept_id:
        return f"dept:{dept_id}"
    identities = scoped_user_identities(user_context)
    return f"user:{identities[0]}" if identities else "none"


def scoped_audit_archives_query(db: Session, request: Request):
    query = db.query(AuditReportArchive)
    context = user_context_from_request(request)
    return query if can_view_all_data(context) else query.filter(AuditReportArchive.scope_key == report_scope_key(context))


def scoped_asset_id_query(db: Session, request: Request) -> list[str]:
    return [row[0] for row in scoped_assets_query(db, request).with_entities(Asset.asset_id).all()]


def users_by_identity(db: Session) -> dict[str, UserDirectory]:
    users = db.query(UserDirectory).all()
    result = {}
    for user in users:
        for key in [user.user_id, user.username, user.email, user.external_id]:
            if key:
                result[str(key)] = user
    return result


def build_audit_workbook(result: dict) -> Workbook:
    workbook = Workbook()
    build_audit_summary_sheet(workbook.active, result)
    build_audit_violations_sheet(workbook.create_sheet("风险明细"), result.get("violations") or [])
    build_audit_responses_sheet(workbook.create_sheet("审计答复"), result.get("responses") or [])
    return workbook


CSV_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def csv_safe(value) -> str:
    """防止 CSV 公式注入：以 = + - @ 等开头的单元格加单引号前缀。"""
    text = "" if value is None else str(value)
    if text.startswith(CSV_FORMULA_PREFIXES):
        return "'" + text
    return text


def csv_response(filename: str, headers: list[str], rows: list[list]) -> PlainTextResponse:
    output = StringIO()
    output.write("\ufeff")
    writer = csv.writer(output)
    writer.writerow([csv_safe(header) for header in headers])
    writer.writerows([[csv_safe(cell) for cell in row] for row in rows])
    return PlainTextResponse(
        output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def date_text(value) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def money(value) -> float:
    return float(value or 0)


def asset_info(asset: Asset) -> str:
    return " / ".join(str(value) for value in [asset.name, asset.brand, asset.model] if value)


def build_audit_summary_sheet(sheet, result: dict) -> None:
    sheet.title = "审计概览"
    rows = [
        ["指标", "值"],
        ["资产总数", result.get("total_assets", 0)],
        ["风险评分", result.get("risk_score", 0)],
        ["风险总数", len(result.get("violations") or [])],
        ["人员风险", (result.get("audit_summary") or {}).get("person", 0)],
        ["资产风险", (result.get("audit_summary") or {}).get("asset", 0)],
    ]
    for row in rows:
        sheet.append(row)
    style_sheet(sheet)


def build_audit_violations_sheet(sheet, violations: list[dict]) -> None:
    sheet.append(["范围", "规则", "等级", "资产ID", "资产编号", "资产信息", "资产名称", "责任人", "部门", "说明", "处理结论", "答复原因", "答复人"])
    for item in violations:
        sheet.append([
            item.get("audit_scope", ""),
            item.get("rule", ""),
            item.get("severity", ""),
            item.get("asset_id", ""),
            item.get("asset_no", ""),
            item.get("asset_info", ""),
            item.get("asset_name", ""),
            item.get("owner_name") or item.get("owner_user_id") or "",
            item.get("dept", ""),
            item.get("message") or item.get("description") or item.get("type") or "",
            item.get("decision", ""),
            item.get("response_reason", ""),
            item.get("responder", ""),
        ])
    style_sheet(sheet)


def build_audit_responses_sheet(sheet, responses: list[dict]) -> None:
    sheet.append(["风险Key", "资产ID", "规则", "范围", "结论", "原因", "答复人", "更新时间"])
    for item in responses:
        sheet.append([
            item.get("violation_key", ""),
            item.get("asset_id", ""),
            item.get("rule_code", ""),
            item.get("audit_scope", ""),
            item.get("decision", ""),
            item.get("reason", ""),
            item.get("responder", ""),
            date_text(item.get("updated_at")),
        ])
    style_sheet(sheet)


def style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def month_windows(count: int) -> list[tuple[str, datetime, datetime]]:
    today = app_now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    starts = []
    current = today
    for _ in range(count - 1):
        starts.append(current)
        current = (current - timedelta(days=1)).replace(day=1)
    starts.append(current)
    starts = list(reversed(starts))
    windows = []
    for start in starts:
        end = (
            start.replace(year=start.year + 1, month=1)
            if start.month == 12
            else start.replace(month=start.month + 1)
        )
        windows.append((start.strftime("%Y-%m"), start.astimezone(timezone.utc), end.astimezone(timezone.utc)))
    return windows
